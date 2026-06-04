import os
import io
import re
import traceback
from collections import defaultdict
from datetime import datetime
from flask import Flask, render_template_string, request, jsonify, redirect, Response
from openpyxl import load_workbook

try:
    import pdfplumber
    TEM_PDFPLUMBER = True
except ImportError:
    TEM_PDFPLUMBER = False

app = Flask(__name__)
app.secret_key = os.environ.get("FLASK_SECRET_KEY", "ifp-dashboard-secret-2026")

# ═════════════════════════════════════════════════════════════
# METAS
# ═════════════════════════════════════════════════════════════
METAS = {
    "matriculas":       120,
    "ticket_medio":     199.0,
    "financeiro_atual": 0.94,
    "frequencia":       0.75,
    "retencao":         0.94,
}
META_MINIMA_BOM = 3

# ═════════════════════════════════════════════════════════════
# FILTROS JINJA2
# ═════════════════════════════════════════════════════════════
def fmt_brl(value):
    try:
        v=float(value); s=f"{v:,.2f}".replace(",","X").replace(".",",").replace("X",".")
        return f"R$ {s}"
    except Exception: return "R$ 0,00"
def fmt_brl0(value):
    try:
        v=float(value); s=f"{v:,.0f}".replace(",",".")
        return f"R$ {s}"
    except Exception: return "R$ 0"
def fmt_pct(value):
    try: return f"{float(value)*100:.1f}%"
    except Exception: return "0,0%"
def fmt_int(value):
    try: return str(int(float(value)))
    except Exception: return "0"
app.jinja_env.filters["brl"]=fmt_brl
app.jinja_env.filters["brl0"]=fmt_brl0
app.jinja_env.filters["pct"]=fmt_pct
app.jinja_env.filters["toint"]=fmt_int

# ═════════════════════════════════════════════════════════════
# CONVERSÃO / DATAS
# ═════════════════════════════════════════════════════════════
def to_float(v, default=0.0):
    if v is None: return default
    if isinstance(v, bool): return default
    if isinstance(v,(int,float)):
        f=float(v); return default if (f!=f) else f
    s=str(v).strip()
    if not s or s in ("-","—","#N/A","#DIV/0!","#VALOR!","N/A","n/a"): return default
    s=s.replace("R$","").replace("%","").replace("\xa0","").replace(" ","").strip()
    if not s: return default
    if re.search(r'\d\.\d{3},',s): s=s.replace(".","").replace(",",".")
    elif "," in s and "." not in s: s=s.replace(",",".")
    elif "," in s and "." in s: s=s.replace(",","")
    s=re.sub(r"[^\d.\-]","",s)
    if not s or s in("-","."): return default
    try: return float(s)
    except ValueError: return default

def parse_data(v):
    if v is None: return None
    if isinstance(v,datetime): return v
    s=str(v).strip()
    for fmt in ("%d/%m/%Y %H:%M:%S","%d/%m/%Y","%Y-%m-%d %H:%M:%S","%Y-%m-%d"):
        try: return datetime.strptime(s,fmt)
        except Exception: pass
    return None

def celula_str(v): return str(v).strip().lower() if v is not None else ""

def norm_nome(nome):
    """Normaliza nome de unidade para casar Excel x VF (remove acento/espaço/caixa)."""
    s = nome.lower().strip()
    s = s.replace("ifpa","ifp")  # IFPA e IFP tratados como prefixo equivalente p/ casar
    for a,b in [("á","a"),("à","a"),("ã","a"),("â","a"),("é","e"),("ê","e"),
                ("í","i"),("ó","o"),("ô","o"),("õ","o"),("ú","u"),("ç","c")]:
        s=s.replace(a,b)
    s=re.sub(r"[^a-z0-9]","",s)
    return s

# ═════════════════════════════════════════════════════════════
# PARSER DO VF (PDF) — fonte OFICIAL dos números
# ═════════════════════════════════════════════════════════════
VF_PATTERN = re.compile(
    r'^\s*\d+\s+(IFPA?\s*-\s*.+?)\s+(\d+)\s+([\d,]+)\s*%\s+R\$\s*([\d.,]+)\s+\d+\s+'
    r'R\$\s*([\d.,]+)\s+R\$\s*([\d.,]+)\s+\d+\s+([\d,]+)\s+'
    r'([\d,]+)%\s*\|\s*([\d,]+)%\s*\|\s*([\d,]+)%\s+'
    r'R\$\s*([\d.,]+)\s+(\d+)\s+(\d+)\s+([\d,]+)\s*%'
)

def parse_vf_pdf(pdf_bytes):
    """Extrai os dados oficiais do VF. Retorna dict {nome_normalizado: {...}}."""
    if not TEM_PDFPLUMBER:
        raise RuntimeError("pdfplumber não instalado. Rode: pip install pdfplumber")
    dados = {}
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        txt = "\n".join(p.extract_text() or "" for p in pdf.pages)
    for linha in txt.split("\n"):
        m = VF_PATTERN.match(linha)
        if not m:
            continue
        nome      = m.group(1).strip()
        matric    = int(m.group(2))
        tpcv1     = to_float(m.group(3))/100.0
        fat_com   = to_float(m.group(4))
        ticket    = to_float(m.group(5))
        media_dia = to_float(m.group(7))
        fin_atual = to_float(m.group(8))/100.0
        fin_30    = to_float(m.group(9))/100.0
        fin_60    = to_float(m.group(10))/100.0
        fat_total = to_float(m.group(11))
        cancel    = int(m.group(12))
        canc_scpc = int(m.group(13))
        dados[norm_nome(nome)] = {
            "nome_vf": nome, "matriculas": matric, "tpcv1": tpcv1,
            "fat_comercial": fat_com, "ticket_medio": ticket,
            "media_diaria": media_dia, "fin_atual": fin_atual,
            "fin_30": fin_30, "fin_60": fin_60, "fat_total": fat_total,
            "cancelados": cancel, "canc_scpc": canc_scpc,
        }
    print(f"[VF] {len(dados)} unidades extraídas do PDF")
    return dados

# ═════════════════════════════════════════════════════════════
# AGRUPAMENTO DE ABAS DO EXCEL
# ═════════════════════════════════════════════════════════════
def tipo_da_aba(nome):
    nl=nome.strip().lower()
    if "(visit" in nl: return "visitas"
    if "(matr" in nl: return "matricula_quitacao"
    if "(frequ" in nl or "(freq" in nl: return "frequencia"
    if "(hist" in nl: return "historico"
    return None

def nome_unidade_da_aba(nome):
    pos=nome.rfind("(")
    return nome[:pos].strip() if pos>0 else nome.strip()

def agrupar_abas(wb):
    grupos={}
    for nome in wb.sheetnames:
        tipo=tipo_da_aba(nome)
        if tipo and tipo!="historico":
            grupos.setdefault(nome_unidade_da_aba(nome),{})[tipo]=wb[nome]
    return grupos

def ler_linhas(ws,max_rows=20000):
    out=[]
    for r in ws.iter_rows(values_only=True):
        out.append(list(r))
        if len(out)>=max_rows: break
    return out

def idx_coluna(header,*nomes):
    for j,h in enumerate(header):
        for nome in nomes:
            if nome in h: return j
    return None

def header_lower(rows): return [celula_str(c) for c in rows[0]] if rows else []

# ═════════════════════════════════════════════════════════════
# PARSERS DO EXCEL (frequência, retenção, alunos — o que o VF não traz)
# ═════════════════════════════════════════════════════════════
def parse_visitas(ws):
    res={"matriculas":0}
    rows=ler_linhas(ws)
    if len(rows)<2: return res
    ci=idx_coluna(header_lower(rows),"status")
    if ci is None: return res
    mat=sum(1 for r in rows[1:] if r and ci<len(r) and celula_str(r[ci]) in ("matrícula","matricula"))
    res["matriculas"]=mat
    return res

def parse_matricula_quitacao(ws):
    res={"ticket_medio":0.0,"fat_total":0.0,"fin_atual":0.0,"valor_atual":0.0,
         "ativos":0,"cancelados":0,"desistentes":0,"nunca_veio":0,"retencao":0.0,"matriculas_mq":0}
    rows=ler_linhas(ws)
    if len(rows)<2: return res
    hdr=header_lower(rows)
    ci_tipo=idx_coluna(hdr,"tipo cobrança","tipo cobranca")
    ci_contrato=idx_coluna(hdr,"contrato")
    ci_status=idx_coluna(hdr,"status contrato","status")
    ci_pago=idx_coluna(hdr,"valor pago")
    ci_valorbase=idx_coluna(hdr,"valor")
    # status únicos
    cstat={}
    for r in rows[1:]:
        if not r: continue
        c=celula_str(r[ci_contrato]) if ci_contrato is not None and ci_contrato<len(r) else ""
        s=celula_str(r[ci_status]) if ci_status is not None and ci_status<len(r) else ""
        if c: cstat[c]=s
    sc=defaultdict(int)
    for s in cstat.values():
        if "ativo" in s or "aprovad" in s or "matrícula" in s or "matricula" in s or "resgate" in s: sc["ativo"]+=1
        elif "nunca" in s: sc["nunca_veio"]+=1
        elif "desist" in s: sc["desistente"]+=1
        elif "cancel" in s: sc["cancelado"]+=1
    res["ativos"]=sc["ativo"]; res["cancelados"]=sc["cancelado"]
    res["desistentes"]=sc["desistente"]; res["nunca_veio"]=sc["nunca_veio"]
    base=sc["ativo"]+sc["cancelado"]+sc["desistente"]
    res["retencao"]=(sc["ativo"]/base) if base>0 else 0.0
    # ticket / faturamento (fallback se VF faltar)
    tk=[]
    for r in rows[1:]:
        if not r or ci_tipo is None or ci_tipo>=len(r): continue
        if celula_str(r[ci_tipo])=="matricula":
            v=to_float(r[ci_valorbase]) if ci_valorbase is not None and ci_valorbase<len(r) else 0
            if v>0: tk.append(v)
    res["ticket_medio"]=(sum(tk)/len(tk)) if tk else 0.0
    res["matriculas_mq"]=len(tk)
    fat=0.0
    for r in rows[1:]:
        if not r or ci_pago is None or ci_pago>=len(r): continue
        fat+=to_float(r[ci_pago])
    res["fat_total"]=fat
    dev=pago=0.0
    for r in rows[1:]:
        if not r or ci_tipo is None or ci_tipo>=len(r): continue
        if celula_str(r[ci_tipo])!="parcela": continue
        vd=to_float(r[ci_valorbase]) if ci_valorbase is not None and ci_valorbase<len(r) else 0
        vp=to_float(r[ci_pago]) if ci_pago is not None and ci_pago<len(r) else 0
        if vd>0: dev+=vd; pago+=vp
    res["fin_atual"]=(pago/dev) if dev>0 else 0.0
    res["valor_atual"]=pago
    return res

def parse_frequencia(ws):
    rows=ler_linhas(ws)
    if len(rows)<2: return 0.0
    hdr=header_lower(rows)
    ci_al=idx_coluna(hdr,"alunos"); ci_pr=idx_coluna(hdr,"presentes","presente")
    ci_fq=idx_coluna(hdr,"frequência","frequencia")
    if ci_al is not None and ci_pr is not None:
        ta=tp=0.0; n=0
        for r in rows[1:]:
            if not r: continue
            a=to_float(r[ci_al]) if ci_al<len(r) else 0
            p=to_float(r[ci_pr]) if ci_pr<len(r) else 0
            if 0<a<=200 and 0<=p<=a: ta+=a; tp+=p; n+=1
        if ta>0 and n>0: return tp/ta
    if ci_fq is not None:
        vals=[]
        for r in rows[1:]:
            if not r or ci_fq>=len(r): continue
            v=to_float(r[ci_fq])
            if 0<v<=1.0: vals.append(v)
        if vals: return sum(vals)/len(vals)
    return 0.0

# ═════════════════════════════════════════════════════════════
# PARSE PRINCIPAL — combina Excel (freq/alunos) + VF (financeiro oficial)
# ═════════════════════════════════════════════════════════════
def parse_tudo(excel_bytes, vf_bytes):
    wb=load_workbook(io.BytesIO(excel_bytes),read_only=True,data_only=True)
    grupos=agrupar_abas(wb)
    if not grupos:
        raise KeyError(f"Nenhuma aba reconhecida no Excel. Abas: {wb.sheetnames[:8]}")
    vf=parse_vf_pdf(vf_bytes)
    if not vf:
        raise ValueError("Não consegui ler os dados do VF (PDF). Verifique se é o arquivo correto.")

    unidades=[]
    for nome_unidade,abas in sorted(grupos.items()):
        u={"nome":nome_unidade,
           "matriculas":0.0,"fat_comercial":0.0,"ticket_medio":0.0,"media_diaria":0.0,"fat_total":0.0,
           "fin_atual":0.0,"fin_30":0.0,"fin_60":0.0,
           "valor_atual":0.0,"valor_30":0.0,"valor_60":0.0,"valor_spc":0.0,"valor_cancelados":0.0,
           "ativos":0.0,"cancelados":0.0,"desistentes":0.0,"nunca_veio":0.0,
           "m1_v1":0.0,"m1_v2":0.0,"frequencia":0.0,"retencao":0.0,
           "tpcv1":0.0,"canc_scpc":0.0,"fonte_vf":False}

        # ── EXCEL: frequência, retenção, alunos ──
        if "matricula_quitacao" in abas:
            d=parse_matricula_quitacao(abas["matricula_quitacao"])
            u["ativos"]=d["ativos"]; u["cancelados"]=d["cancelados"]
            u["desistentes"]=d["desistentes"]; u["nunca_veio"]=d["nunca_veio"]
            u["retencao"]=d["retencao"]
            # fallback (caso VF não tenha a unidade)
            u["ticket_medio"]=d["ticket_medio"]; u["fat_total"]=d["fat_total"]
            u["fin_atual"]=d["fin_atual"]; u["valor_atual"]=d["valor_atual"]
            u["matriculas"]=d["matriculas_mq"]
        if "visitas" in abas:
            mv=parse_visitas(abas["visitas"])["matriculas"]
            if mv>0: u["matriculas"]=mv
        if "frequencia" in abas:
            f=parse_frequencia(abas["frequencia"])
            if f>0: u["frequencia"]=f

        # ── VF: SOBRESCREVE com os valores oficiais (matrículas, ticket,
        #        faturamento, financeiro atual/30/60, cancelados) ──
        chave=norm_nome(nome_unidade)
        if chave in vf:
            v=vf[chave]
            u["matriculas"]=v["matriculas"]
            u["ticket_medio"]=v["ticket_medio"]
            u["fat_comercial"]=v["fat_comercial"]
            u["media_diaria"]=v["media_diaria"]
            u["fat_total"]=v["fat_total"]
            u["fin_atual"]=v["fin_atual"]
            u["fin_30"]=v["fin_30"]
            u["fin_60"]=v["fin_60"]
            u["cancelados"]=v["cancelados"]
            u["tpcv1"]=v["tpcv1"]
            u["canc_scpc"]=v["canc_scpc"]
            u["fonte_vf"]=True
            # valores R$ de cobrança 30/60 (proporção do faturamento)
            u["valor_atual"]=v["fat_comercial"]*v["fin_atual"]
            u["valor_30"]=v["fat_comercial"]*v["fin_30"]
            u["valor_60"]=v["fat_comercial"]*v["fin_60"]
        else:
            print(f"[AVISO] '{nome_unidade}' não encontrada no VF (chave={chave})")

        unidades.append(calcular_score(u))
    return unidades

# ═════════════════════════════════════════════════════════════
# SCORE + RANKING
# ═════════════════════════════════════════════════════════════
def metas_efetivas(u):
    """Retorna as metas aplicáveis à unidade. Se houver metas decendiais salvas
       para esta unidade, usa-as onde houver equivalente (Matrícula, Atual→cobrança,
       Ticket da parcela→ticket). Freq e Retenção permanecem nas metas padrão."""
    m = dict(METAS)
    dm = _dec["metas"].get(u["nome"], {}) if "_dec" in globals() else {}
    if dm:
        mat = to_float(dm.get("matricula",{}).get("meta_mes",0))
        if mat>0: m["matriculas"]=mat
        atual = to_float(dm.get("atual",{}).get("meta_mes",0))
        # "Atual" é uma % (ex: 94,00). Aceita tanto 94 quanto 0.94
        if atual>0: m["financeiro_atual"]= atual/100.0 if atual>1.5 else atual
        tk = to_float(dm.get("ticket_parcela",{}).get("valor",0))
        if tk>0: m["ticket_medio"]=tk
    return m

def calcular_score(u):
    M = metas_efetivas(u)
    score=0
    if u["matriculas"]>=M["matriculas"]: score+=1
    if u["ticket_medio"]>=M["ticket_medio"]: score+=1
    if u["fin_atual"]>=M["financeiro_atual"]: score+=1
    if u["frequencia"]>=M["frequencia"]: score+=1
    if u["retencao"]>=M["retencao"]: score+=1
    u["score"]=score
    u["status"]="bom" if score>=META_MINIMA_BOM else "ruim"
    def nrm(val,meta,cap=2.0): return min(val/meta,cap) if meta else 0
    pts=(nrm(u["matriculas"],M["matriculas"])*20+nrm(u["ticket_medio"],M["ticket_medio"])*20+
         nrm(u["fin_atual"],M["financeiro_atual"])*20+nrm(u["frequencia"],M["frequencia"])*20+
         nrm(u["retencao"],M["retencao"])*20)/2.0
    u["rank_score"]=round(pts,1)
    def bar(val,mx): return 0 if mx==0 else max(0.0,min(100.0,val/mx*100))
    def cls(val,meta):
        if val>=meta: return "verde"
        if val>=meta*0.85: return "amarelo"
        return "vermelho"
    u["indicators"]=[
        {"label":"Matrículas","display":str(int(u["matriculas"])),"bar":bar(u["matriculas"],max(M["matriculas"]*1.6,200)),
         "cls":cls(u["matriculas"],M["matriculas"]),"meta":f"Meta: {int(M['matriculas'])}"},
        {"label":"Ticket Médio","display":fmt_brl0(u["ticket_medio"]),"bar":bar(u["ticket_medio"],max(M["ticket_medio"]*2,400)),
         "cls":cls(u["ticket_medio"],M["ticket_medio"]),"meta":f"Meta: {fmt_brl0(M['ticket_medio'])}"},
        {"label":"Cobrança Atual","display":fmt_pct(u["fin_atual"]),"bar":u["fin_atual"]*100,
         "cls":cls(u["fin_atual"],M["financeiro_atual"]),"meta":f"Meta: {fmt_pct(M['financeiro_atual'])}"},
        {"label":"Frequência","display":fmt_pct(u["frequencia"]),"bar":u["frequencia"]*100,
         "cls":cls(u["frequencia"],M["frequencia"]),"meta":f"Meta: {fmt_pct(M['frequencia'])}"},
        {"label":"Retenção","display":fmt_pct(u["retencao"]),"bar":u["retencao"]*100,
         "cls":cls(u["retencao"],M["retencao"]),"meta":f"Meta: {fmt_pct(M['retencao'])}"},
    ]
    return u

def ranquear(unidades):
    ordenadas=sorted(unidades,key=lambda u:(u["rank_score"],u["score"],u["fat_total"]),reverse=True)
    for i,u in enumerate(ordenadas): u["posicao"]=i+1
    return ordenadas

def ranquear_por(unidades, chave, reverse=True, somente_vf=True, desempate=None, desempate_reverse=True):
    """Ranking genérico por um campo. reverse=True → maior é melhor.
       somente_vf filtra unidades que casaram com o VF (têm o dado oficial).
       desempate: campo usado como 2º critério quando o 1º empata."""
    base = [u for u in unidades if (u.get("fonte_vf") or not somente_vf)]
    def chave_ordem(u):
        primario = u.get(chave, 0)
        # ordena sempre crescente internamente, invertendo o sinal quando reverse
        p = -primario if reverse else primario
        if desempate is not None:
            d = u.get(desempate, 0)
            s = -d if desempate_reverse else d
            return (p, s)
        return (p,)
    ordenadas = sorted(base, key=chave_ordem)
    for i,u in enumerate(ordenadas): u["pos_crit"]=i+1
    return ordenadas

# ═════════════════════════════════════════════════════════════
# CACHE
# ═════════════════════════════════════════════════════════════
_cache={"data":None,"periodo":"","filename":"","wb_bytes":None}
def get_cached_data(): return _cache["data"],_cache["periodo"],_cache["filename"]
def set_cached_data(data,periodo="",filename="",wb_bytes=None):
    _cache.update({"data":data,"periodo":periodo,"filename":filename,"wb_bytes":wb_bytes})

# ═════════════════════════════════════════════════════════════
# PERÍODO DECENDIAL — metas por unidade + resultados por decêndio
# ═════════════════════════════════════════════════════════════
# Linhas do formulário (na ordem da imagem de exemplo)
LINHAS_DEC = [
    {"id":"matricula",      "label":"Matrícula",                 "tipo":"valor"},
    {"id":"proj_comercial", "label":"Projeção Comercial",        "tipo":"valor"},
    {"id":"atual",          "label":"Atual",                     "tipo":"pct"},
    {"id":"d30",            "label":"30 Dias",                   "tipo":"pct"},
    {"id":"d60",            "label":"60 Dias",                   "tipo":"pct"},
    {"id":"d90",            "label":"90 Dias",                   "tipo":"pct"},
    {"id":"scpc",           "label":"SCPC",                      "tipo":"valor"},
    {"id":"cancelamentos",  "label":"Cancelamentos",             "tipo":"valor"},
    {"id":"parc_antecip",   "label":"Parcelas Antecipadas",      "tipo":"valor"},
    {"id":"atraso90",       "label":"Atrasado + 90 Dias",        "tipo":"valor"},
    {"id":"produtos",       "label":"Produtos",                  "tipo":"valor"},
    {"id":"ticket_parcela", "label":"Ticket da parcela (Carteira Somente Mês Atual)", "tipo":"simples"},
    {"id":"qtd_parcelas",   "label":"Quantidade de Parcelas no Início do Mês (Somente Mês Atual)", "tipo":"simples"},
]
# Decêndios do mês
DECENDIOS = [
    {"id":"d1",  "label":"Dia 1 ao dia 10"},
    {"id":"d2",  "label":"Dia 11 ao dia 21"},
    {"id":"d3",  "label":"Dia 22 ao último dia"},
]
# Campos editáveis por linha (mês anterior, meta mês, quantidade, valor)
CAMPOS_DEC = ["mes_anterior","meta_mes","quantidade","valor"]

# Armazenamento em memória:
#   _metas_dec[nome_unidade][linha_id] = {"mes_anterior":x,"meta_mes":y,"quantidade":z,"valor":w}
#   _result_dec[nome_unidade][decendio_id] = {...comparação VF x meta...}
_dec = {"metas":{}, "resultados":{}, "unidades":[]}

def dec_unidades():
    """Lista de nomes de unidades — vem do Fechamento Mensal se já carregado,
       senão da lista salva no próprio decendial."""
    data,_,_ = get_cached_data()
    if data:
        nomes=[u["nome"] for u in data]
        _dec["unidades"]=nomes
        return nomes
    return _dec.get("unidades",[])

def metas_da_unidade(nome):
    return _dec["metas"].get(nome, {})

def total_recebido(nome):
    """Total Recebido = soma da coluna VALOR de todas as linhas."""
    m = _dec["metas"].get(nome, {})
    return sum(to_float(m.get(l["id"],{}).get("valor",0)) for l in LINHAS_DEC)


# ═════════════════════════════════════════════════════════════
# DEBUG
# ═════════════════════════════════════════════════════════════
@app.route("/debug")
def debug_index():
    if not _cache["wb_bytes"]: return "<h2>Nenhuma planilha</h2>",404
    wb=load_workbook(io.BytesIO(_cache["wb_bytes"]),read_only=True,data_only=True)
    grupos=agrupar_abas(wb)
    links="".join(f'<li>{k} — {list(v.keys())}</li>' for k,v in sorted(grupos.items()))
    return f"<h2>Unidades ({len(grupos)})</h2><ul style='font-family:monospace'>{links}</ul>"


# ═════════════════════════════════════════════════════════════
# CSS COMPARTILHADO — identidade IFP refinada
# ═════════════════════════════════════════════════════════════
CSS = r"""
@import url('https://fonts.googleapis.com/css2?family=Sora:wght@400;500;600;700;800&family=Plus+Jakarta+Sans:wght@400;500;600;700;800&display=swap');
:root{
  --red:#c0021c;--red-dk:#8b0014;--red-br:#e30a2a;--red-lt:rgba(192,2,28,.06);
  --blue:#16205e;--blue-2:#1f2d7a;--blue-lt:#3949ab;--blue-soft:#eef0fb;
  --bg:#eef1f7;--bg-2:#f6f8fc;--card:#fff;--borda:#e2e6f2;--txt:#141a36;--txt2:#646c8c;
  --verde:#1b7a3d;--verde-lt:#2fad5a;--verde-bg:#e7f6ec;
  --amber:#c2620a;--amber-lt:#f59020;--amber-bg:#fff4e6;
  --rose:#c0021c;--rose-bg:#fdebed;
  --ouro:#d4af37;--ouro-2:#f4cf52;--prata:#9aa3b2;--prata-2:#c7cedb;--bronze:#cd7f32;--bronze-2:#e0a96d;
  --r:18px;--r-sm:12px;
  --shadow:0 4px 18px rgba(22,32,94,.08);--shadow-lg:0 14px 44px rgba(22,32,94,.16);
  --shadow-red:0 10px 30px rgba(192,2,28,.22);
}
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:'Plus Jakarta Sans',sans-serif;background:
  radial-gradient(1200px 500px at 85% -10%,rgba(192,2,28,.05),transparent),
  radial-gradient(1000px 400px at 0% 0%,rgba(22,32,94,.05),transparent),var(--bg);
  color:var(--txt);min-height:100vh}
h1,h2,h3,.cnome,.tc-val,.podium-nome,.detail-hero h2{font-family:'Sora',sans-serif}

/* HEADER */
.hdr{background:linear-gradient(115deg,var(--red-dk) 0%,var(--red) 55%,var(--red-br) 100%);
  color:#fff;position:relative;overflow:hidden}
.hdr::after{content:"";position:absolute;inset:0;background:
  radial-gradient(600px 200px at 90% 120%,rgba(255,255,255,.12),transparent);pointer-events:none}
.hdr-in{max-width:1600px;margin:0 auto;display:flex;align-items:center;justify-content:space-between;padding:16px 26px;gap:12px;flex-wrap:wrap;position:relative}
.logo-wrap{display:flex;align-items:center;gap:15px}
.logo-img{width:54px;height:54px;background:#fff;border-radius:14px;display:flex;align-items:center;justify-content:center;box-shadow:0 6px 18px rgba(0,0,0,.18)}
.logo-img span{font-family:'Sora';font-size:1.15rem;font-weight:800;color:var(--red);letter-spacing:-1px}
.logo-txt h1{font-size:1.2rem;font-weight:800;color:#fff;line-height:1.15;letter-spacing:-.3px}
.logo-txt p{font-size:.74rem;opacity:.82;margin-top:3px;font-weight:500}
.hdr-badge{background:rgba(255,255,255,.16);border:1.5px solid rgba(255,255,255,.4);border-radius:30px;padding:7px 18px;font-size:.82rem;font-weight:700;color:#fff;backdrop-filter:blur(6px)}

/* NAV */
.nav{background:var(--blue);padding:0 26px;box-shadow:0 3px 14px rgba(22,32,94,.2)}
.nav-in{max-width:1600px;margin:0 auto;display:flex;align-items:center;gap:4px;flex-wrap:wrap}
.nav a{padding:14px 20px;color:rgba(255,255,255,.62);text-decoration:none;font-size:.87rem;font-weight:700;border-bottom:3px solid transparent;margin-bottom:-1px;transition:.18s;display:flex;align-items:center;gap:7px}
.nav a:hover{color:#fff;background:rgba(255,255,255,.06)}
.nav a.ativo{color:#fff;border-bottom-color:var(--red-br)}

/* TOOLBAR */
.toolbar{background:var(--card);border-bottom:1px solid var(--borda);padding:12px 26px}
.toolbar-in{max-width:1600px;margin:0 auto;display:flex;align-items:center;gap:12px;flex-wrap:wrap}
.toolbar label{font-size:.82rem;font-weight:700;color:var(--txt2)}
.toolbar input[type=file]{font-size:.78rem;padding:7px 10px;border:1.5px solid var(--borda);border-radius:10px;background:var(--bg-2);color:var(--txt)}
.toolbar input[type=file]::-webkit-file-upload-button{background:var(--red);color:#fff;border:none;border-radius:7px;padding:5px 12px;font-size:.76rem;font-weight:700;cursor:pointer;margin-right:8px}
.btn{border:none;border-radius:11px;font:inherit;font-weight:700;cursor:pointer;padding:9px 20px;font-size:.84rem;text-decoration:none;display:inline-flex;align-items:center;gap:7px;transition:.18s}
.btn-red{background:var(--red);color:#fff;box-shadow:0 4px 14px rgba(192,2,28,.28)}.btn-red:hover{background:var(--red-dk);transform:translateY(-1px)}
.btn-blue{background:var(--blue);color:#fff}.btn-blue:hover{background:var(--blue-2);transform:translateY(-1px)}
.btn-ghost{background:var(--blue-soft);color:var(--blue);}.btn-ghost:hover{background:#e0e4f8}
.smsg{font-size:.8rem;font-weight:700}.smsg.ok{color:var(--verde)}.smsg.err{color:var(--rose)}

/* FILTROS */
.filtros{background:var(--card);border-bottom:1px solid var(--borda);padding:12px 26px}
.filtros-in{max-width:1600px;margin:0 auto;display:flex;align-items:center;gap:10px;flex-wrap:wrap}
.flabel{font-size:.8rem;font-weight:800;color:var(--blue);letter-spacing:.02em}
.fbtn{padding:7px 17px;border-radius:30px;border:1.5px solid var(--borda);background:var(--bg-2);color:var(--txt2);font-size:.8rem;font-weight:700;cursor:pointer;transition:.16s}
.fbtn:hover{border-color:var(--red);color:var(--red)}
.fbtn.ativo{background:var(--blue);color:#fff;border-color:var(--blue)}
.search{padding:8px 16px;border-radius:30px;border:1.5px solid var(--borda);background:var(--bg-2);font-size:.82rem;outline:none;width:220px;font-family:inherit;transition:.16s}
.search:focus{border-color:var(--red);box-shadow:0 0 0 3px var(--red-lt)}

/* TOTAIS */
.totais{max-width:1600px;margin:22px auto 0;padding:0 26px}
.totais-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(200px,1fr));gap:14px}
.tc{background:var(--card);border-radius:var(--r);padding:18px 18px 16px;box-shadow:var(--shadow);position:relative;overflow:hidden;border:1px solid var(--borda);transition:.2s;display:flex;flex-direction:column;min-height:118px;cursor:pointer}
.tc:hover{box-shadow:var(--shadow-lg);transform:translateY(-2px)}
.tc::before{content:"";position:absolute;top:0;left:0;right:0;height:5px;background:linear-gradient(90deg,var(--blue),var(--blue-lt))}
.tc-ico{position:absolute;top:15px;right:16px;font-size:1.15rem;opacity:.22;line-height:1}
.tc-lbl{font-size:.66rem;font-weight:800;color:var(--txt2);text-transform:uppercase;letter-spacing:.06em;padding-right:26px;line-height:1.25}
.tc-val{font-size:clamp(1.35rem,2.4vw,1.85rem);font-weight:800;color:var(--blue);line-height:1.05;margin-top:auto;padding-top:8px;letter-spacing:-.6px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.tc.aberto .tc-val{white-space:normal;overflow:visible;text-overflow:clip;font-size:1.25rem;word-break:break-word}
.tc.aberto{box-shadow:var(--shadow-lg)}
.tc-hint{position:absolute;bottom:10px;right:14px;font-size:.6rem;font-weight:700;color:var(--txt2);opacity:0;transition:.16s}
.tc:hover .tc-hint{opacity:.6}
.tc-sub{font-size:.7rem;color:var(--txt2);margin-top:5px;font-weight:600;line-height:1.2}
.tc.r::before{background:linear-gradient(90deg,var(--red-dk),var(--red-br))}.tc.r .tc-val{color:var(--red)}
.tc.g::before{background:linear-gradient(90deg,var(--verde),var(--verde-lt))}.tc.g .tc-val{color:var(--verde)}
.tc.x::before{background:linear-gradient(90deg,#9a0012,#e53935)}.tc.x .tc-val{color:var(--rose)}

/* CARDS */
.cards{max-width:1600px;margin:22px auto 60px;padding:0 26px}
.cards-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(352px,1fr));gap:16px}
.ucard{background:var(--card);border-radius:var(--r);box-shadow:var(--shadow);overflow:hidden;border:1px solid var(--borda);transition:.22s;cursor:pointer;position:relative;display:block;text-decoration:none;color:inherit}
.ucard:hover{box-shadow:var(--shadow-lg);transform:translateY(-4px)}
.ucard:hover .chd::after{opacity:1;transform:translateX(0)}
.chd{padding:15px 16px;display:flex;align-items:center;justify-content:space-between;gap:8px;position:relative;overflow:hidden}
.chd::before{content:"";position:absolute;inset:0;background:radial-gradient(400px 120px at 110% -30%,rgba(255,255,255,.18),transparent)}
.chd::after{content:"➜";position:absolute;right:14px;top:50%;transform:translate(8px,-50%);margin-top:0;color:rgba(255,255,255,.9);font-size:1rem;opacity:0;transition:.2s}
.chd.bom{background:linear-gradient(135deg,#155e30,#1f9048);color:#fff}
.chd.ruim{background:linear-gradient(135deg,var(--red-dk),var(--red));color:#fff}
.cnome{font-size:.95rem;font-weight:800;flex:1;letter-spacing:-.2px;position:relative;z-index:1}
.sbadge{background:rgba(255,255,255,.22);border:1px solid rgba(255,255,255,.4);border-radius:30px;padding:4px 11px;font-size:.7rem;font-weight:800;white-space:nowrap;position:relative;z-index:1;backdrop-filter:blur(4px)}
.cbody{padding:13px 14px;display:flex;flex-direction:column;gap:5px}
.ind{display:flex;align-items:center;gap:9px;padding:6px 9px;border-radius:9px}
.ind.verde{background:var(--verde-bg)}.ind.amarelo{background:var(--amber-bg)}.ind.vermelho{background:var(--rose-bg)}
.ind-lbl{font-size:.74rem;font-weight:700;color:var(--txt2);flex:1}
.ind-meta{font-size:.6rem;color:var(--txt2);opacity:.75;font-weight:600}.ind-val{font-size:.85rem;font-weight:800}
.ind.verde .ind-val{color:var(--verde)}.ind.amarelo .ind-val{color:var(--amber)}.ind.vermelho .ind-val{color:var(--rose)}
.barwrap{width:54px;height:6px;background:#e3e7f1;border-radius:4px;overflow:hidden;flex-shrink:0}
.bar{height:100%;border-radius:4px}.bar.verde{background:linear-gradient(90deg,var(--verde),var(--verde-lt))}.bar.amarelo{background:linear-gradient(90deg,var(--amber),var(--amber-lt))}.bar.vermelho{background:linear-gradient(90deg,#9a0012,#e53935)}
.sec-title{font-size:.64rem;font-weight:800;color:var(--txt2);text-transform:uppercase;letter-spacing:.06em;padding:8px 0 3px;border-top:1px dashed var(--borda);margin-top:5px;display:flex;align-items:center;gap:6px}
.sec-tag{font-size:.55rem;font-weight:800;padding:2px 7px;border-radius:8px;text-transform:none;letter-spacing:0}
.sec-tag.mq{background:var(--blue-soft);color:var(--blue)}.sec-tag.vf{background:var(--amber-bg);color:var(--amber)}
.row3{display:grid;grid-template-columns:1fr 1fr 1fr;gap:6px}.row4{display:grid;grid-template-columns:repeat(4,1fr);gap:6px}
.box{background:var(--bg-2);border-radius:9px;padding:7px 7px;text-align:center;border:1px solid var(--borda)}
.box-lbl{font-size:.57rem;font-weight:800;color:var(--txt2);text-transform:uppercase;letter-spacing:.03em}
.box-val{font-size:.9rem;font-weight:800;color:var(--blue);margin-top:3px}
.box.ok .box-val{color:var(--verde)}.box.al .box-val{color:var(--red)}
.fd3{display:grid;grid-template-columns:1fr 1fr 1fr;gap:6px}
.fd{background:var(--bg-2);border-radius:9px;padding:6px 8px;border:1px solid var(--borda)}
.fd-l{font-size:.55rem;font-weight:800;color:var(--txt2);text-transform:uppercase}
.fd-p{font-size:.92rem;font-weight:800;margin-top:2px}.fd-v{font-size:.64rem;color:var(--txt2);margin-top:1px;font-weight:600}
.fd.ok .fd-p{color:var(--verde)}.fd.am .fd-p{color:var(--amber)}.fd.bad .fd-p{color:var(--rose)}
.fat-total{display:flex;align-items:center;justify-content:space-between;padding:9px 12px;border-radius:11px;background:linear-gradient(135deg,var(--blue-soft),#dfe3f7);margin-top:5px;border:1px solid #d2d8f0}
.fat-lbl{font-size:.73rem;font-weight:800;color:var(--blue)}.fat-val{font-size:.98rem;font-weight:800;color:var(--blue)}
.card-cta{display:flex;align-items:center;justify-content:center;gap:6px;margin-top:9px;padding:8px 0;border-radius:11px;background:var(--blue-soft);color:var(--blue);font-size:.76rem;font-weight:800;transition:.16s}
.ucard:hover .card-cta{background:var(--red);color:#fff}

/* LANDING / upload */
.landing{max-width:680px;margin:54px auto;padding:0 24px}
.land-card{background:var(--card);border-radius:24px;box-shadow:var(--shadow-lg);overflow:hidden}
.land-top{background:linear-gradient(135deg,var(--red-dk),var(--red) 70%,var(--red-br));padding:40px 38px 30px;text-align:center;position:relative;overflow:hidden}
.land-top::after{content:"";position:absolute;inset:0;background:radial-gradient(500px 200px at 80% 130%,rgba(255,255,255,.14),transparent)}
.land-top .lico{font-size:3.6rem;margin-bottom:10px;position:relative}
.land-top h2{font-size:1.6rem;font-weight:800;color:#fff;margin-bottom:7px;position:relative}
.land-top p{color:rgba(255,255,255,.9);font-size:.92rem;line-height:1.55;position:relative}
.land-body{padding:32px 38px}.land-body h3{font-size:1.05rem;font-weight:800;color:var(--blue);margin-bottom:16px}
.uprow{display:flex;flex-direction:column;gap:14px}
.upbox{border:2px dashed var(--borda);border-radius:14px;padding:18px;background:var(--bg-2);display:flex;align-items:center;gap:14px;transition:.18s}
.upbox.ok{border-color:var(--verde-lt);background:var(--verde-bg)}
.upbox-ico{font-size:1.9rem;flex-shrink:0}
.upbox-txt{flex:1}.upbox-txt b{font-size:.9rem;color:var(--blue);display:block}
.upbox-txt span{font-size:.74rem;color:var(--txt2)}
.upbox input[type=file]{font-size:.74rem;max-width:165px}
.upbox input[type=file]::-webkit-file-upload-button{background:var(--red);color:#fff;border:none;border-radius:7px;padding:6px 10px;font-size:.72rem;font-weight:700;cursor:pointer;margin-right:6px}
.land-btn{display:block;width:100%;margin-top:20px;padding:14px;border-radius:13px;background:var(--red);color:#fff;border:none;font:inherit;font-size:.96rem;font-weight:800;cursor:pointer;box-shadow:var(--shadow-red);transition:.18s}
.land-btn:hover{background:var(--red-dk);transform:translateY(-2px)}
.land-hints{margin-top:20px;display:flex;flex-direction:column;gap:9px}
.hint{display:flex;align-items:flex-start;gap:10px;padding:11px 13px;border-radius:11px;background:var(--amber-bg);font-size:.78rem;color:#8a5200;font-weight:600}
.hint-icon{font-size:1.1rem}

/* DETALHE DA UNIDADE */
.detail{max-width:1100px;margin:24px auto 60px;padding:0 26px}
.detail-back{display:inline-flex;align-items:center;gap:7px;font-size:.82rem;font-weight:700;color:var(--blue);text-decoration:none;margin-bottom:16px;padding:7px 14px;border-radius:30px;background:var(--card);box-shadow:var(--shadow);transition:.16s}
.detail-back:hover{background:var(--blue);color:#fff;transform:translateX(-3px)}
.detail-hero{border-radius:var(--r);overflow:hidden;box-shadow:var(--shadow-lg);margin-bottom:20px}
.detail-hero-top{padding:26px 28px;color:#fff;display:flex;align-items:center;justify-content:space-between;gap:16px;flex-wrap:wrap;position:relative;overflow:hidden}
.detail-hero-top::after{content:"";position:absolute;inset:0;background:radial-gradient(500px 200px at 95% 130%,rgba(255,255,255,.15),transparent)}
.detail-hero-top.bom{background:linear-gradient(135deg,#155e30,#1f9048)}
.detail-hero-top.ruim{background:linear-gradient(135deg,var(--red-dk),var(--red))}
.detail-hero h2{font-size:1.8rem;font-weight:800;position:relative;letter-spacing:-.5px}
.detail-hero .sub{font-size:.84rem;opacity:.88;margin-top:4px;position:relative;font-weight:500}
.detail-status{text-align:center;position:relative}
.detail-status .big{font-size:2.6rem;font-weight:800;font-family:'Sora';line-height:1}
.detail-status .lbl{font-size:.72rem;font-weight:800;text-transform:uppercase;letter-spacing:.08em;opacity:.9;margin-top:3px}
.detail-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(220px,1fr));gap:14px;margin-bottom:18px}
.dcard{background:var(--card);border-radius:var(--r);box-shadow:var(--shadow);padding:18px 20px;border:1px solid var(--borda)}
.dcard h4{font-size:.7rem;font-weight:800;color:var(--txt2);text-transform:uppercase;letter-spacing:.06em;margin-bottom:12px;display:flex;align-items:center;gap:7px}
.dind{display:flex;align-items:center;justify-content:space-between;padding:9px 0;border-bottom:1px solid var(--borda)}
.dind:last-child{border-bottom:none}
.dind-l{font-size:.82rem;font-weight:600;color:var(--txt)}
.dind-r{display:flex;align-items:center;gap:10px}
.dind-v{font-size:1rem;font-weight:800}.dind-v.verde{color:var(--verde)}.dind-v.amarelo{color:var(--amber)}.dind-v.vermelho{color:var(--rose)}
.dbar{width:70px;height:7px;background:#e3e7f1;border-radius:5px;overflow:hidden}
.dmeta{font-size:.62rem;color:var(--txt2);font-weight:600}
.dbig-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(130px,1fr));gap:12px}
.dbig{background:var(--bg-2);border-radius:12px;padding:14px;text-align:center;border:1px solid var(--borda)}
.dbig-v{font-size:1.4rem;font-weight:800;color:var(--blue);font-family:'Sora'}
.dbig-v.g{color:var(--verde)}.dbig-v.r{color:var(--red)}
.dbig-l{font-size:.62rem;font-weight:800;color:var(--txt2);text-transform:uppercase;margin-top:3px}
.detail-actions{display:flex;gap:12px;flex-wrap:wrap;margin-top:6px}

/* RANKING */
.rank-wrap{max-width:1120px;margin:26px auto 60px;padding:0 26px}
.podium{display:grid;grid-template-columns:1fr 1.18fr 1fr;gap:16px;align-items:end;margin-bottom:30px}
.podium-card{border-radius:20px;padding:22px 16px;text-align:center;color:#fff;box-shadow:var(--shadow-lg);position:relative;overflow:hidden;transition:.2s}
.podium-card:hover{transform:translateY(-4px)}
.podium-card::after{content:"";position:absolute;inset:0;background:radial-gradient(300px 120px at 50% -20%,rgba(255,255,255,.25),transparent)}
.podium-card.p1{background:linear-gradient(165deg,var(--ouro-2),var(--ouro));padding-top:34px}
.podium-card.p2{background:linear-gradient(165deg,var(--prata-2),var(--prata))}
.podium-card.p3{background:linear-gradient(165deg,var(--bronze-2),var(--bronze))}
.podium-medal{font-size:2.7rem;line-height:1;margin-bottom:8px;position:relative;filter:drop-shadow(0 3px 6px rgba(0,0,0,.2))}
.podium-pos{font-size:.7rem;font-weight:800;text-transform:uppercase;letter-spacing:.09em;opacity:.95;position:relative}
.podium-nome{font-size:1.1rem;font-weight:800;margin:5px 0 10px;line-height:1.15;position:relative}
.podium-score{font-size:2.1rem;font-weight:800;font-family:'Sora';position:relative}
.podium-score small{font-size:.7rem;font-weight:700;opacity:.85}
.podium-meta{font-size:.72rem;font-weight:700;margin-top:7px;background:rgba(255,255,255,.25);border-radius:12px;padding:4px 10px;display:inline-block;position:relative}
.rank-table{background:var(--card);border-radius:var(--r);box-shadow:var(--shadow);overflow:hidden;border:1px solid var(--borda)}
.rank-row{display:grid;grid-template-columns:56px 1fr repeat(5,minmax(58px,86px)) 76px 70px;align-items:center;gap:8px;padding:13px 18px;border-bottom:1px solid var(--borda);font-size:.82rem;transition:.14s;text-decoration:none;color:inherit}
.rank-row.head{background:var(--blue);color:#fff;font-weight:800;font-size:.68rem;text-transform:uppercase;letter-spacing:.04em}
.rank-row:last-child{border-bottom:none}
.rank-row:not(.head):hover{background:var(--blue-soft)}
.rank-row.top:not(.head){background:linear-gradient(90deg,rgba(212,175,55,.07),transparent)}
.rank-pos{font-weight:800;font-size:1.05rem;color:var(--blue);text-align:center}
.rank-pos.top1{color:var(--ouro)}.rank-pos.top2{color:var(--prata)}.rank-pos.top3{color:var(--bronze)}
.rank-nome{font-weight:800}
.rank-cell{text-align:center;font-weight:700}.rank-cell.g{color:var(--verde)}.rank-cell.r{color:var(--rose)}
.rank-score-cell{text-align:center;font-weight:800;color:var(--blue);font-size:.98rem;font-family:'Sora'}
.rank-badge{font-size:.6rem;font-weight:800;padding:3px 9px;border-radius:30px}
.rank-badge.bom{background:var(--verde-bg);color:var(--verde)}.rank-badge.ruim{background:var(--rose-bg);color:var(--rose)}
@media(max-width:920px){.rank-row{grid-template-columns:42px 1fr 66px 66px;font-size:.74rem}.rank-hide{display:none}}
@media(max-width:640px){.cards-grid{grid-template-columns:1fr}.totais-grid{grid-template-columns:repeat(2,1fr)}.row3,.fd3{grid-template-columns:1fr 1fr}.row4{grid-template-columns:repeat(2,1fr)}.podium{grid-template-columns:1fr}.podium-card.p1{order:-1}}
@media print{.hdr,.nav,.toolbar,.filtros,.totais,.card-cta,.btn,.btn-blue,.detail-back,.detail-actions{display:none!important}.cards,.rank-wrap,.detail{margin:0;padding:0}.cards-grid{grid-template-columns:repeat(2,1fr)}.ucard{break-inside:avoid;box-shadow:none;border:1px solid #ccc}body{background:#fff}}
"""

# ═════════════════════════════════════════════════════════════
# TEMPLATE DASHBOARD
# ═════════════════════════════════════════════════════════════
TEMPLATE = r"""
<!DOCTYPE html><html lang="pt-br"><head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>IFP – Dashboard</title><style>__CSS__</style></head><body>
<div class="hdr"><div class="hdr-in">
  <div class="logo-wrap"><div class="logo-img"><span>IFP</span></div>
  <div class="logo-txt"><h1>Dashboard de Fechamento Mensal</h1><p>Instituto de Formação Profissional</p></div></div>
  {% if periodo %}<div class="hdr-badge">📅 {{ periodo }}</div>{% endif %}
</div></div>
<div class="nav"><div class="nav-in"><a href="/fechamento" class="ativo">📊 Painel</a><a href="/ranking">🏆 Ranking Geral</a><a href="/ranking/comercial">💰 Ranking Comercial</a><a href="/ranking/tpcv1">📉 Ranking TPCv1</a></div></div>
{% if unidades %}
<div class="toolbar"><div class="toolbar-in">
  <form method="POST" action="/upload" enctype="multipart/form-data" style="display:flex;align-items:center;gap:8px;flex-wrap:wrap">
    <label>📊 Excel:</label><input type="file" name="planilha" accept=".xlsx,.xlsm" required>
    <label>📄 VF:</label><input type="file" name="vf" accept=".pdf" required>
    <button type="submit" class="btn btn-red">⬆ Atualizar</button>
  </form>
  <a href="/pdf/todas" class="btn btn-blue" target="_blank">⬇ PDF Geral</a>
  {% if msg %}<span class="smsg {{ 'ok' if msg_ok else 'err' }}">{{ msg }}</span>{% endif %}
</div></div>
<div class="filtros"><div class="filtros-in">
  <span class="flabel">Filtrar:</span>
  <button class="fbtn ativo" onclick="filtrar('todos',this)">Todos ({{ unidades|length }})</button>
  <button class="fbtn" onclick="filtrar('bom',this)" style="color:#1b7a3d;border-color:#9bdcb0">✅ Bom ({{ unidades|selectattr('status','eq','bom')|list|length }})</button>
  <button class="fbtn" onclick="filtrar('ruim',this)" style="color:#c0021c;border-color:#f0a9b1">❌ Ruim ({{ unidades|selectattr('status','eq','ruim')|list|length }})</button>
  <input type="text" class="search" placeholder="🔍 Buscar unidade..." oninput="buscar(this.value)">
</div></div>
<div class="totais"><div class="totais-grid">
  <div class="tc"><span class="tc-ico">🎓</span><span class="tc-lbl">Matrículas Totais</span><span class="tc-val">{{ totais.matriculas|toint }}</span><span class="tc-sub">{{ unidades|length }} unidades</span></div>
  <div class="tc r"><span class="tc-ico">👥</span><span class="tc-lbl">Alunos Ativos</span><span class="tc-val">{{ totais.ativos|toint }}</span><span class="tc-sub">Retenção {{ totais.retencao_str }}</span></div>
  <div class="tc"><span class="tc-ico">💰</span><span class="tc-lbl">Fat. Comercial</span><span class="tc-val">{{ totais.fat_comercial|brl0 }}</span><span class="tc-sub">Ticket {{ totais.ticket_str }}</span></div>
  <div class="tc"><span class="tc-ico">🏦</span><span class="tc-lbl">Fat. Total</span><span class="tc-val">{{ totais.fat_total|brl0 }}</span><span class="tc-sub">Carteira total</span></div>
  <div class="tc"><span class="tc-ico">📋</span><span class="tc-lbl">Cobr. Atual Média</span><span class="tc-val">{{ totais.fin_atual_str }}</span><span class="tc-sub">Meta 94%</span></div>
  <div class="tc"><span class="tc-ico">📈</span><span class="tc-lbl">Frequência Média</span><span class="tc-val">{{ totais.freq_str }}</span><span class="tc-sub">Meta 75%</span></div>
  <div class="tc g"><span class="tc-ico">✅</span><span class="tc-lbl">Unidades Boas</span><span class="tc-val">{{ unidades|selectattr('status','eq','bom')|list|length }}</span><span class="tc-sub">3 ou mais metas</span></div>
  <div class="tc x"><span class="tc-ico">❌</span><span class="tc-lbl">Unidades Ruins</span><span class="tc-val">{{ unidades|selectattr('status','eq','ruim')|list|length }}</span><span class="tc-sub">Menos de 3 metas</span></div>
</div></div>
<div class="cards"><div class="cards-grid" id="cards-grid">
{% for u in unidades %}
<a href="/unidade/{{ loop.index0 }}" class="ucard" data-status="{{ u.status }}" data-nome="{{ u.nome|lower }}">
  <div class="chd {{ u.status }}">
    <span class="cnome">{{ u.nome }}</span>
    <span class="sbadge">{%- if u.status=='bom' %}✅ BOM{%- else %}❌ RUIM{%- endif -%}&nbsp;· {{ u.score }}/5</span>
  </div>
  <div class="cbody">
    {% for ind in u.indicators %}
    <div class="ind {{ ind.cls }}"><span class="ind-lbl">{{ ind.label }}</span><span class="ind-meta">{{ ind.meta }}</span>
      <span class="ind-val">{{ ind.display }}</span>
      <div class="barwrap"><div class="bar {{ ind.cls }}" style="width:{{ ind.bar|round(1) }}%"></div></div></div>
    {% endfor %}
    <span class="sec-title">📋 Cobrança <span class="sec-tag vf">VF Oficial</span></span>
    <div class="fd3">
      {% set fa = 'ok' if u.fin_atual>=0.94 else ('am' if u.fin_atual>=0.80 else 'bad') %}
      <div class="fd {{ fa }}"><div class="fd-l">Atual</div><div class="fd-p">{{ u.fin_atual|pct }}</div><div class="fd-v">{{ u.valor_atual|brl0 }}</div></div>
      <div class="fd"><div class="fd-l">30 dias</div><div class="fd-p">{{ u.fin_30|pct }}</div><div class="fd-v">{{ u.valor_30|brl0 }}</div></div>
      <div class="fd"><div class="fd-l">60 dias</div><div class="fd-p">{{ u.fin_60|pct }}</div><div class="fd-v">{{ u.valor_60|brl0 }}</div></div>
    </div>
    <span class="sec-title">💰 Faturamento <span class="sec-tag vf">VF Oficial</span></span>
    <div class="row3">
      <div class="box"><div class="box-lbl">Comercial</div><div class="box-val">{{ u.fat_comercial|brl0 }}</div></div>
      <div class="box"><div class="box-lbl">Ticket</div><div class="box-val">{{ u.ticket_medio|brl0 }}</div></div>
      <div class="box"><div class="box-lbl">Matríc.</div><div class="box-val">{{ u.matriculas|toint }}</div></div>
    </div>
    <span class="sec-title">👥 Alunos <span class="sec-tag mq">Matrícula e Quitação</span></span>
    <div class="row4">
      <div class="box ok"><div class="box-lbl">Ativos</div><div class="box-val">{{ u.ativos|toint }}</div></div>
      <div class="box {{ 'al' if u.cancelados>20 else '' }}"><div class="box-lbl">Cancel.</div><div class="box-val">{{ u.cancelados|toint }}</div></div>
      <div class="box {{ 'al' if u.desistentes>50 else '' }}"><div class="box-lbl">Desist.</div><div class="box-val">{{ u.desistentes|toint }}</div></div>
      <div class="box {{ 'al' if u.nunca_veio>30 else '' }}"><div class="box-lbl">N.Veio</div><div class="box-val">{{ u.nunca_veio|toint }}</div></div>
    </div>
    <span class="sec-title">🚫 Cancelamentos <span class="sec-tag vf">VF Oficial</span></span>
    <div class="row3">
      <div class="box {{ 'al' if u.cancelados>20 else '' }}"><div class="box-lbl">Cancelamentos</div><div class="box-val">{{ u.cancelados|toint }}</div></div>
      <div class="box {{ 'al' if u.canc_scpc>0 else '' }}"><div class="box-lbl">Canc. SCPC</div><div class="box-val">{{ u.canc_scpc|toint }}</div></div>
      <div class="box"><div class="box-lbl">TPCv1</div><div class="box-val">{{ u.tpcv1|pct }}</div></div>
    </div>
    <div class="fat-total"><span class="fat-lbl">Faturamento Total</span><span class="fat-val">{{ u.fat_total|brl }}</span></div>
    <div class="card-cta">Ver detalhes da unidade ➜</div>
  </div>
</a>
{% endfor %}
</div></div>
{% else %}
<div class="landing"><div class="land-card">
  <div class="land-top"><div class="lico">📊</div><h2>Dashboard IFP</h2>
    <p>Envie os DOIS arquivos do fechamento mensal para visualizar os indicadores.</p></div>
  <div class="land-body"><h3>Carregar arquivos</h3>
    <form method="POST" action="/upload" enctype="multipart/form-data">
      <div class="uprow">
        <div class="upbox" id="box-excel"><span class="upbox-ico">📊</span>
          <div class="upbox-txt"><b>1. Planilha Excel</b><span>RC - Núcleo de Inteligência (.xlsx)</span></div>
          <input type="file" name="planilha" id="in-excel" accept=".xlsx,.xlsm" required onchange="mark('box-excel',this)"></div>
        <div class="upbox" id="box-vf"><span class="upbox-ico">📄</span>
          <div class="upbox-txt"><b>2. Relatório VF</b><span>VF - Fechamento (.pdf) — oficial 30/60 dias</span></div>
          <input type="file" name="vf" id="in-vf" accept=".pdf" required onchange="mark('box-vf',this)"></div>
      </div>
      <button type="submit" class="land-btn">⬆ Carregar e Processar</button>
      {% if msg %}<p style="margin-top:12px;text-align:center;font-size:.85rem;font-weight:700;color:{{ '#c0021c' if not msg_ok else '#1b7a3d' }}">{{ msg }}</p>{% endif %}
    </form>
    <div class="land-hints">
      <div class="hint"><span class="hint-icon">⚠️</span><span>Os <strong>dois arquivos são obrigatórios</strong>. O VF fornece os valores oficiais de Cobrança Atual/30/60 e faturamento.</span></div>
      <div class="hint"><span class="hint-icon">📋</span><span>O Excel fornece frequência, retenção e situação dos alunos.</span></div>
    </div>
  </div>
</div></div>
{% endif %}
<script>
function mark(id,inp){var b=document.getElementById(id);if(inp.files[0]){b.classList.add('ok');b.querySelector('span:last-child').textContent='✓ '+inp.files[0].name}}
var fa='todos';
function filtrar(s,b){fa=s;document.querySelectorAll('.fbtn').forEach(function(x){x.classList.remove('ativo')});if(b)b.classList.add('ativo');ap(document.querySelector('.search')?document.querySelector('.search').value.toLowerCase():'')}
function buscar(q){ap(q.toLowerCase())}
function ap(q){q=q||'';document.querySelectorAll('.ucard').forEach(function(c){c.style.display=((fa==='todos'||c.dataset.status===fa)&&(!q||c.dataset.nome.includes(q)))?'':'none'})}
// Totalizadores: clicar expande o valor inteiro quando ele está cortado
document.querySelectorAll('.tc').forEach(function(card){
  var val=card.querySelector('.tc-val');
  // adiciona dica só quando o texto está truncado
  function cortado(){return val.scrollWidth>val.clientWidth+1;}
  var hint=document.createElement('span');hint.className='tc-hint';hint.textContent='ver completo';
  card.appendChild(hint);
  function atualizaHint(){hint.style.display=(cortado()||card.classList.contains('aberto'))?'':'none';}
  atualizaHint();
  card.addEventListener('click',function(){
    card.classList.toggle('aberto');
    hint.textContent=card.classList.contains('aberto')?'recolher':'ver completo';
    atualizaHint();
  });
  window.addEventListener('resize',atualizaHint);
});
</script>
</body></html>
"""

# ═════════════════════════════════════════════════════════════
# TEMPLATE DETALHE DA UNIDADE
# ═════════════════════════════════════════════════════════════
UNIDADE_TEMPLATE = r"""
<!DOCTYPE html><html lang="pt-br"><head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>IFP – {{ u.nome }}</title><style>__CSS__</style></head><body>
<div class="hdr"><div class="hdr-in">
  <div class="logo-wrap"><div class="logo-img"><span>IFP</span></div>
  <div class="logo-txt"><h1>{{ u.nome }}</h1><p>Detalhamento da unidade</p></div></div>
  {% if periodo %}<div class="hdr-badge">📅 {{ periodo }}</div>{% endif %}
</div></div>
<div class="nav"><div class="nav-in"><a href="/fechamento">📊 Painel</a><a href="/ranking">🏆 Ranking Geral</a><a href="/ranking/comercial">💰 Ranking Comercial</a><a href="/ranking/tpcv1">📉 Ranking TPCv1</a></div></div>
<div class="detail">
  <a href="/fechamento" class="detail-back">⬅ Voltar ao painel</a>
  <div class="detail-hero">
    <div class="detail-hero-top {{ u.status }}">
      <div><h2>{{ u.nome }}</h2><div class="sub">{%- if u.status=='bom' %}Unidade com bom desempenho{%- else %}Unidade abaixo das metas{%- endif -%} · Posição #{{ posicao }} no ranking</div></div>
      <div class="detail-status"><div class="big">{{ u.score }}/5</div><div class="lbl">{%- if u.status=='bom' %}✅ BOM{%- else %}❌ RUIM{%- endif -%}</div></div>
    </div>
  </div>
  <div class="detail-grid">
    <div class="dcard">
      <h4>🎯 Indicadores vs Metas</h4>
      {% for ind in u.indicators %}
      <div class="dind">
        <span class="dind-l">{{ ind.label }}</span>
        <div class="dind-r">
          <span class="dmeta">{{ ind.meta }}</span>
          <div class="dbar"><div class="bar {{ ind.cls }}" style="width:{{ ind.bar|round(1) }}%"></div></div>
          <span class="dind-v {{ ind.cls }}">{{ ind.display }}</span>
        </div>
      </div>
      {% endfor %}
    </div>
    <div class="dcard">
      <h4>📋 Cobrança <span class="sec-tag vf">VF Oficial</span></h4>
      <div class="dbig-grid">
        <div class="dbig"><div class="dbig-v {{ 'g' if u.fin_atual>=0.94 else 'r' }}">{{ u.fin_atual|pct }}</div><div class="dbig-l">Atual</div></div>
        <div class="dbig"><div class="dbig-v">{{ u.fin_30|pct }}</div><div class="dbig-l">30 dias</div></div>
        <div class="dbig"><div class="dbig-v">{{ u.fin_60|pct }}</div><div class="dbig-l">60 dias</div></div>
      </div>
    </div>
  </div>
  <div class="detail-grid">
    <div class="dcard">
      <h4>💰 Faturamento <span class="sec-tag vf">VF Oficial</span></h4>
      <div class="dbig-grid">
        <div class="dbig"><div class="dbig-v">{{ u.fat_comercial|brl0 }}</div><div class="dbig-l">Comercial</div></div>
        <div class="dbig"><div class="dbig-v">{{ u.ticket_medio|brl0 }}</div><div class="dbig-l">Ticket Médio</div></div>
        <div class="dbig"><div class="dbig-v">{{ u.matriculas|toint }}</div><div class="dbig-l">Matrículas</div></div>
        <div class="dbig"><div class="dbig-v g">{{ u.fat_total|brl0 }}</div><div class="dbig-l">Fat. Total</div></div>
      </div>
    </div>
    <div class="dcard">
      <h4>👥 Alunos <span class="sec-tag mq">Matrícula e Quitação</span></h4>
      <div class="dbig-grid">
        <div class="dbig"><div class="dbig-v g">{{ u.ativos|toint }}</div><div class="dbig-l">Ativos</div></div>
        <div class="dbig"><div class="dbig-v r">{{ u.cancelados|toint }}</div><div class="dbig-l">Cancelados</div></div>
        <div class="dbig"><div class="dbig-v r">{{ u.desistentes|toint }}</div><div class="dbig-l">Desistentes</div></div>
        <div class="dbig"><div class="dbig-v">{{ u.nunca_veio|toint }}</div><div class="dbig-l">Nunca Veio</div></div>
      </div>
    </div>
  </div>
  <div class="detail-grid">
    <div class="dcard">
      <h4>🚫 Cancelamentos <span class="sec-tag vf">VF Oficial</span></h4>
      <div class="dbig-grid">
        <div class="dbig"><div class="dbig-v r">{{ u.cancelados|toint }}</div><div class="dbig-l">Cancelamentos</div></div>
        <div class="dbig"><div class="dbig-v {{ 'r' if u.canc_scpc>0 else '' }}">{{ u.canc_scpc|toint }}</div><div class="dbig-l">Canc. SCPC</div></div>
        <div class="dbig"><div class="dbig-v">{{ u.tpcv1|pct }}</div><div class="dbig-l">TPCv1</div></div>
        <div class="dbig"><div class="dbig-v">{{ u.media_diaria|round(1) }}</div><div class="dbig-l">Média Diária</div></div>
      </div>
    </div>
  </div>
  <div class="detail-actions">
    <a href="/pdf/unidade/{{ idx }}" target="_blank" class="btn btn-red">⬇ Baixar PDF desta unidade</a>
    <a href="/ranking" class="btn btn-ghost">🏆 Ver ranking geral</a>
  </div>
</div>
</body></html>
"""

# ═════════════════════════════════════════════════════════════
# TEMPLATE RANKING
# ═════════════════════════════════════════════════════════════
RANKING_TEMPLATE = r"""
<!DOCTYPE html><html lang="pt-br"><head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>IFP – {{ titulo }}</title><style>__CSS__</style></head><body>
<div class="hdr"><div class="hdr-in">
  <div class="logo-wrap"><div class="logo-img"><span>IFP</span></div>
  <div class="logo-txt"><h1>{{ titulo }}</h1><p>Instituto de Formação Profissional</p></div></div>
  {% if periodo %}<div class="hdr-badge">📅 {{ periodo }}</div>{% endif %}
</div></div>
<div class="nav"><div class="nav-in">
  <a href="/fechamento">📊 Painel</a>
  <a href="/ranking" class="{{ 'ativo' if modo=='geral' else '' }}">🏆 Ranking Geral</a>
  <a href="/ranking/comercial" class="{{ 'ativo' if modo=='comercial' else '' }}">💰 Ranking Comercial</a>
  <a href="/ranking/tpcv1" class="{{ 'ativo' if modo=='tpcv1' else '' }}">📉 Ranking TPCv1</a>
</div></div>
{% if ranking %}
<div class="toolbar"><div class="toolbar-in">
  <span class="flabel">{{ subtitulo }}</span>
  <div style="flex:1"></div>
  <a href="{{ pdf_url }}" class="btn btn-red" target="_blank">⬇ Baixar este Ranking em PDF</a>
</div></div>
<div class="rank-wrap">
  <div class="podium">
    {% if ranking|length>1 %}<a href="/unidade/{{ ranking[1].idx_orig }}" class="podium-card p2" style="text-decoration:none"><div class="podium-medal">🥈</div><div class="podium-pos">2º Lugar</div><div class="podium-nome">{{ ranking[1].nome }}</div><div class="podium-score">{{ ranking[1].metric_str }}</div><div class="podium-meta">{{ ranking[1].metric_sub }}</div></a>{% else %}<div></div>{% endif %}
    <a href="/unidade/{{ ranking[0].idx_orig }}" class="podium-card p1" style="text-decoration:none"><div class="podium-medal">🥇</div><div class="podium-pos">1º Lugar</div><div class="podium-nome">{{ ranking[0].nome }}</div><div class="podium-score">{{ ranking[0].metric_str }}</div><div class="podium-meta">{{ ranking[0].metric_sub }}</div></a>
    {% if ranking|length>2 %}<a href="/unidade/{{ ranking[2].idx_orig }}" class="podium-card p3" style="text-decoration:none"><div class="podium-medal">🥉</div><div class="podium-pos">3º Lugar</div><div class="podium-nome">{{ ranking[2].nome }}</div><div class="podium-score">{{ ranking[2].metric_str }}</div><div class="podium-meta">{{ ranking[2].metric_sub }}</div></a>{% else %}<div></div>{% endif %}
  </div>
  <div class="rank-table">
    {% if modo=='geral' %}
    <div class="rank-row head"><div>#</div><div>Unidade</div><div class="rank-hide">Matríc.</div><div class="rank-hide">Ticket</div><div class="rank-hide">Cobr.</div><div class="rank-hide">Freq.</div><div class="rank-hide">Retenç.</div><div>Pontos</div><div class="rank-hide">Status</div></div>
    {% for u in ranking %}
    <a href="/unidade/{{ u.idx_orig }}" class="rank-row {{ 'top' if u.posicao<=3 else '' }}">
      <div class="rank-pos {% if u.posicao==1 %}top1{% elif u.posicao==2 %}top2{% elif u.posicao==3 %}top3{% endif %}">{% if u.posicao==1 %}🥇{% elif u.posicao==2 %}🥈{% elif u.posicao==3 %}🥉{% else %}{{ u.posicao }}{% endif %}</div>
      <div class="rank-nome">{{ u.nome }}</div>
      <div class="rank-cell rank-hide {{ 'g' if u.matriculas>=120 else 'r' }}">{{ u.matriculas|toint }}</div>
      <div class="rank-cell rank-hide {{ 'g' if u.ticket_medio>=199 else 'r' }}">{{ u.ticket_medio|brl0 }}</div>
      <div class="rank-cell rank-hide {{ 'g' if u.fin_atual>=0.94 else 'r' }}">{{ u.fin_atual|pct }}</div>
      <div class="rank-cell rank-hide {{ 'g' if u.frequencia>=0.75 else 'r' }}">{{ u.frequencia|pct }}</div>
      <div class="rank-cell rank-hide {{ 'g' if u.retencao>=0.94 else 'r' }}">{{ u.retencao|pct }}</div>
      <div class="rank-score-cell">{{ u.rank_score }}</div>
      <div class="rank-hide"><span class="rank-badge {{ u.status }}">{{ 'BOM' if u.status=='bom' else 'RUIM' }}</span></div>
    </a>
    {% endfor %}
    {% elif modo=='comercial' %}
    <div class="rank-row head" style="grid-template-columns:56px 1fr minmax(120px,1fr) minmax(80px,120px) minmax(80px,120px)"><div>#</div><div>Unidade</div><div>Fat. Comercial</div><div class="rank-hide">Ticket</div><div class="rank-hide">Matríc.</div></div>
    {% for u in ranking %}
    <a href="/unidade/{{ u.idx_orig }}" class="rank-row {{ 'top' if u.pos_crit<=3 else '' }}" style="grid-template-columns:56px 1fr minmax(120px,1fr) minmax(80px,120px) minmax(80px,120px)">
      <div class="rank-pos {% if u.pos_crit==1 %}top1{% elif u.pos_crit==2 %}top2{% elif u.pos_crit==3 %}top3{% endif %}">{% if u.pos_crit==1 %}🥇{% elif u.pos_crit==2 %}🥈{% elif u.pos_crit==3 %}🥉{% else %}{{ u.pos_crit }}{% endif %}</div>
      <div class="rank-nome">{{ u.nome }}</div>
      <div class="rank-score-cell" style="text-align:left">{{ u.fat_comercial|brl }}</div>
      <div class="rank-cell rank-hide">{{ u.ticket_medio|brl0 }}</div>
      <div class="rank-cell rank-hide">{{ u.matriculas|toint }}</div>
    </a>
    {% endfor %}
    {% else %}
    <div class="rank-row head" style="grid-template-columns:56px 1fr minmax(100px,1fr) minmax(80px,120px) minmax(80px,130px)"><div>#</div><div>Unidade</div><div>TPCv1</div><div class="rank-hide">Canc. SCPC</div><div class="rank-hide">Fat. Comercial</div></div>
    {% for u in ranking %}
    <a href="/unidade/{{ u.idx_orig }}" class="rank-row {{ 'top' if u.pos_crit<=3 else '' }}" style="grid-template-columns:56px 1fr minmax(100px,1fr) minmax(80px,120px) minmax(80px,130px)">
      <div class="rank-pos {% if u.pos_crit==1 %}top1{% elif u.pos_crit==2 %}top2{% elif u.pos_crit==3 %}top3{% endif %}">{% if u.pos_crit==1 %}🥇{% elif u.pos_crit==2 %}🥈{% elif u.pos_crit==3 %}🥉{% else %}{{ u.pos_crit }}{% endif %}</div>
      <div class="rank-nome">{{ u.nome }}</div>
      <div class="rank-score-cell" style="text-align:left">{{ u.tpcv1|pct }}</div>
      <div class="rank-cell rank-hide {{ 'r' if u.canc_scpc>0 else 'g' }}">{{ u.canc_scpc|toint }}</div>
      <div class="rank-cell rank-hide">{{ u.fat_comercial|brl0 }}</div>
    </a>
    {% endfor %}
    {% endif %}
  </div>
  {% if modo=='tpcv1' %}<p style="text-align:center;color:var(--txt2);font-size:.78rem;margin-top:14px;font-weight:600">📉 No TPCv1, <strong>quanto menor melhor</strong>. Em caso de empate (ex: várias unidades em 0%), o desempate é pelo <strong>maior Faturamento Comercial</strong>.</p>{% endif %}
</div>
{% else %}
<div class="landing"><div class="land-card"><div class="land-top"><div class="lico">🏆</div><h2>Sem dados</h2><p>Carregue os arquivos no Painel para ver o ranking.</p></div>
<div class="land-body"><a href="/fechamento" class="land-btn" style="text-align:center;text-decoration:none">⬅ Ir para o Painel</a></div></div></div>
{% endif %}
</body></html>
"""

# ═════════════════════════════════════════════════════════════
# TEMPLATES PDF
# ═════════════════════════════════════════════════════════════
PDF_TEMPLATE = r"""
<!DOCTYPE html><html lang="pt-br"><head><meta charset="UTF-8"><title>IFP – {{ titulo }}</title>
<link href="https://fonts.googleapis.com/css2?family=Sora:wght@600;700;800&family=Plus+Jakarta+Sans:wght@400;600;700;800&display=swap" rel="stylesheet">
<style>
*{box-sizing:border-box;margin:0;padding:0}body{font-family:'Plus Jakarta Sans',sans-serif;background:#fff;color:#141a36;font-size:11px}
h2,.bx-v,.fat-v{font-family:'Sora'}
.ph{background:linear-gradient(120deg,#8b0014,#c0021c 60%,#e30a2a);color:#fff;padding:18px 24px;display:flex;align-items:center;justify-content:space-between;margin-bottom:18px}
.ph-l{display:flex;align-items:center;gap:12px}.ph-logo{width:42px;height:42px;background:#fff;border-radius:10px;display:flex;align-items:center;justify-content:center;font-family:'Sora';font-weight:800;color:#c0021c;font-size:1rem}
.ph-r h2{font-size:1rem;font-weight:800;text-align:right}.ph-r p{font-size:.68rem;opacity:.85;text-align:right;margin-top:2px}
.no-print{text-align:center;margin-bottom:14px}.no-print button{border:none;border-radius:10px;padding:10px 26px;font-size:.88rem;font-weight:700;cursor:pointer;margin:0 4px}
.btn-p{background:#c0021c;color:#fff}.btn-c{background:#16205e;color:#fff}
.grid{display:grid;grid-template-columns:repeat(2,1fr);gap:12px;padding:0 18px 18px}
.ucard{border:1.5px solid #e2e6f2;border-radius:13px;overflow:hidden;break-inside:avoid}
.chd{padding:11px 13px;display:flex;align-items:center;justify-content:space-between}
.chd.bom{background:linear-gradient(135deg,#155e30,#1f9048);color:#fff}.chd.ruim{background:linear-gradient(135deg,#8b0014,#c0021c);color:#fff}
.cnome{font-size:.88rem;font-weight:800}.sbadge{font-size:.66rem;font-weight:800;background:rgba(255,255,255,.22);border:1px solid rgba(255,255,255,.35);border-radius:30px;padding:3px 9px}
.cbody{padding:10px 12px;display:flex;flex-direction:column;gap:5px}
.ind{display:flex;align-items:center;gap:6px;padding:5px 8px;border-radius:7px}
.ind.verde{background:#e7f6ec}.ind.amarelo{background:#fff4e6}.ind.vermelho{background:#fdebed}
.ind-lbl{font-size:.68rem;font-weight:700;color:#646c8c;flex:1}.ind-meta{font-size:.56rem;color:#646c8c;opacity:.7}.ind-val{font-size:.78rem;font-weight:800}
.ind.verde .ind-val{color:#1b7a3d}.ind.amarelo .ind-val{color:#c2620a}.ind.vermelho .ind-val{color:#c0021c}
.barwrap{width:44px;height:5px;background:#e3e7f1;border-radius:3px;overflow:hidden}.bar{height:100%;border-radius:3px}
.bar.verde{background:#2fad5a}.bar.amarelo{background:#f59020}.bar.vermelho{background:#e53935}
.sec{font-size:.58rem;font-weight:800;color:#646c8c;text-transform:uppercase;padding:5px 0 2px;border-top:1px dashed #e2e6f2;margin-top:3px}
.r3{display:grid;grid-template-columns:1fr 1fr 1fr;gap:5px}.r4{display:grid;grid-template-columns:repeat(4,1fr);gap:5px}
.bx{background:#f6f8fc;border-radius:7px;padding:5px 6px;text-align:center;border:1px solid #e2e6f2}.bx-l{font-size:.54rem;font-weight:800;color:#646c8c;text-transform:uppercase}.bx-v{font-size:.84rem;font-weight:800;color:#16205e;margin-top:1px}
.bx.ok .bx-v{color:#1b7a3d}.bx.al .bx-v{color:#c0021c}
.fat{display:flex;align-items:center;justify-content:space-between;padding:6px 9px;border-radius:9px;background:linear-gradient(135deg,#eef0fb,#dfe3f7);margin-top:3px}.fat-l{font-size:.68rem;font-weight:800;color:#16205e}.fat-v{font-size:.88rem;font-weight:800;color:#16205e}
.rod{text-align:center;color:#9aa3b2;font-size:.64rem;padding:10px 0 16px;border-top:1px solid #eee;margin:0 18px}
@media print{body{-webkit-print-color-adjust:exact;print-color-adjust:exact}.no-print{display:none}}
</style></head><body>
<div class="ph"><div class="ph-l"><div class="ph-logo">IFP</div></div><div class="ph-r"><h2>{{ titulo }}</h2><p>Instituto de Formação Profissional &nbsp;|&nbsp; {{ periodo }}</p></div></div>
<div class="no-print"><button class="btn-p" onclick="window.print()">🖨️ Imprimir / Salvar PDF</button><button class="btn-c" onclick="window.close()">✕ Fechar</button></div>
<div class="grid">
{% for u in unidades %}
<div class="ucard"><div class="chd {{ u.status }}"><span class="cnome">{{ u.nome }}</span><span class="sbadge">{%- if u.status=='bom' %}✅ BOM{%- else %}❌ RUIM{%- endif -%}&nbsp;· {{ u.score }}/5</span></div>
<div class="cbody">
{% for ind in u.indicators %}<div class="ind {{ ind.cls }}"><span class="ind-lbl">{{ ind.label }}</span><span class="ind-meta">{{ ind.meta }}</span><span class="ind-val">{{ ind.display }}</span><div class="barwrap"><div class="bar {{ ind.cls }}" style="width:{{ ind.bar|round(1) }}%"></div></div></div>{% endfor %}
<span class="sec">Cobrança (VF Oficial)</span><div class="r3"><div class="bx"><div class="bx-l">Atual</div><div class="bx-v">{{ u.fin_atual|pct }}</div></div><div class="bx"><div class="bx-l">30 dias</div><div class="bx-v">{{ u.fin_30|pct }}</div></div><div class="bx"><div class="bx-l">60 dias</div><div class="bx-v">{{ u.fin_60|pct }}</div></div></div>
<span class="sec">Faturamento (VF Oficial)</span><div class="r3"><div class="bx"><div class="bx-l">Comercial</div><div class="bx-v">{{ u.fat_comercial|brl0 }}</div></div><div class="bx"><div class="bx-l">Ticket</div><div class="bx-v">{{ u.ticket_medio|brl0 }}</div></div><div class="bx"><div class="bx-l">Matrículas</div><div class="bx-v">{{ u.matriculas|toint }}</div></div></div>
<span class="sec">Alunos</span><div class="r4"><div class="bx ok"><div class="bx-l">Ativos</div><div class="bx-v">{{ u.ativos|toint }}</div></div><div class="bx {{ 'al' if u.cancelados>20 else '' }}"><div class="bx-l">Cancel.</div><div class="bx-v">{{ u.cancelados|toint }}</div></div><div class="bx {{ 'al' if u.desistentes>50 else '' }}"><div class="bx-l">Desist.</div><div class="bx-v">{{ u.desistentes|toint }}</div></div><div class="bx {{ 'al' if u.nunca_veio>30 else '' }}"><div class="bx-l">N.Veio</div><div class="bx-v">{{ u.nunca_veio|toint }}</div></div></div>
<span class="sec">Cancelamentos (VF Oficial)</span><div class="r3"><div class="bx {{ 'al' if u.cancelados>20 else '' }}"><div class="bx-l">Cancelamentos</div><div class="bx-v">{{ u.cancelados|toint }}</div></div><div class="bx {{ 'al' if u.canc_scpc>0 else '' }}"><div class="bx-l">Canc. SCPC</div><div class="bx-v">{{ u.canc_scpc|toint }}</div></div><div class="bx"><div class="bx-l">TPCv1</div><div class="bx-v">{{ u.tpcv1|pct }}</div></div></div>
<div class="fat"><span class="fat-l">Faturamento Total</span><span class="fat-v">{{ u.fat_total|brl }}</span></div>
</div></div>
{% endfor %}
</div>
<div class="rod">Gerado pelo sistema IFP Dashboard &nbsp;|&nbsp; {{ periodo }}</div>
{% if auto_print %}<script>window.addEventListener('load',function(){setTimeout(function(){window.print()},800)});</script>{% endif %}
</body></html>
"""

PDF_RANKING = r"""
<!DOCTYPE html><html lang="pt-br"><head><meta charset="UTF-8"><title>IFP – Ranking</title>
<link href="https://fonts.googleapis.com/css2?family=Sora:wght@600;700;800&family=Plus+Jakarta+Sans:wght@400;600;700;800&display=swap" rel="stylesheet">
<style>
*{box-sizing:border-box;margin:0;padding:0}body{font-family:'Plus Jakarta Sans',sans-serif;background:#fff;color:#141a36;font-size:11px}
h2,.sc,.nm{font-family:'Sora'}
.ph{background:linear-gradient(120deg,#8b0014,#c0021c 60%,#e30a2a);color:#fff;padding:18px 24px;display:flex;align-items:center;justify-content:space-between;margin-bottom:18px}
.ph-logo{width:42px;height:42px;background:#fff;border-radius:10px;display:flex;align-items:center;justify-content:center;font-family:'Sora';font-weight:800;color:#c0021c}
.ph-r h2{font-size:1rem;font-weight:800;text-align:right}.ph-r p{font-size:.68rem;opacity:.85;text-align:right;margin-top:2px}
.no-print{text-align:center;margin-bottom:14px}.no-print button{border:none;border-radius:10px;padding:10px 26px;font-size:.88rem;font-weight:700;cursor:pointer;margin:0 4px}
.btn-p{background:#c0021c;color:#fff}.btn-c{background:#16205e;color:#fff}
.podium{display:grid;grid-template-columns:1fr 1.18fr 1fr;gap:14px;align-items:end;max-width:780px;margin:0 auto 24px;padding:0 18px}
.pc{border-radius:16px;padding:16px 12px;text-align:center;color:#fff;box-shadow:0 8px 20px rgba(0,0,0,.12)}
.pc.p1{background:linear-gradient(165deg,#f4cf52,#d4af37);padding-top:24px}.pc.p2{background:linear-gradient(165deg,#c7cedb,#9aa3b2)}.pc.p3{background:linear-gradient(165deg,#e0a96d,#cd7f32)}
.pc .m{font-size:2.2rem}.pc .pos{font-size:.65rem;font-weight:800;text-transform:uppercase}.pc .nm{font-size:1rem;font-weight:800;margin:3px 0 6px}.pc .sc{font-size:1.7rem;font-weight:800}
table{border-collapse:collapse;width:calc(100% - 36px);margin:0 18px 18px;font-size:10.5px}
th,td{border:1px solid #e2e6f2;padding:7px 8px;text-align:center}
th{background:#16205e;color:#fff;font-size:.62rem;text-transform:uppercase}
td.nm{text-align:left;font-weight:800}td.pos{font-weight:800;color:#16205e}td.g{color:#1b7a3d;font-weight:700}td.r{color:#c0021c;font-weight:700}td.sc{font-weight:800;color:#16205e;font-family:'Sora'}
tr:nth-child(even){background:#f6f8fc}
.rod{text-align:center;color:#9aa3b2;font-size:.64rem;padding:10px 0 16px;border-top:1px solid #eee;margin:0 18px}
@media print{body{-webkit-print-color-adjust:exact;print-color-adjust:exact}.no-print{display:none}}
</style></head><body>
<div class="ph"><div class="ph-logo">IFP</div><div class="ph-r"><h2>{{ titulo }}</h2><p>Instituto de Formação Profissional &nbsp;|&nbsp; {{ periodo }}</p></div></div>
<div class="no-print"><button class="btn-p" onclick="window.print()">🖨️ Imprimir / Salvar PDF</button><button class="btn-c" onclick="window.close()">✕ Fechar</button></div>
<div class="podium">
  {% if ranking|length>1 %}<div class="pc p2"><div class="m">🥈</div><div class="pos">2º Lugar</div><div class="nm">{{ ranking[1].nome }}</div><div class="sc">{{ ranking[1].metric_str }}</div></div>{% else %}<div></div>{% endif %}
  <div class="pc p1"><div class="m">🥇</div><div class="pos">1º Lugar</div><div class="nm">{{ ranking[0].nome }}</div><div class="sc">{{ ranking[0].metric_str }}</div></div>
  {% if ranking|length>2 %}<div class="pc p3"><div class="m">🥉</div><div class="pos">3º Lugar</div><div class="nm">{{ ranking[2].nome }}</div><div class="sc">{{ ranking[2].metric_str }}</div></div>{% else %}<div></div>{% endif %}
</div>
{% if modo=='geral' %}
<table>
<tr><th>#</th><th style="text-align:left">Unidade</th><th>Matríc.</th><th>Ticket</th><th>Cobr.Atual</th><th>30d</th><th>60d</th><th>Freq.</th><th>Retenç.</th><th>Pontos</th><th>Status</th></tr>
{% for u in ranking %}
<tr>
  <td class="pos">{% if u.posicao==1 %}🥇{% elif u.posicao==2 %}🥈{% elif u.posicao==3 %}🥉{% else %}{{ u.posicao }}{% endif %}</td>
  <td class="nm">{{ u.nome }}</td>
  <td class="{{ 'g' if u.matriculas>=120 else 'r' }}">{{ u.matriculas|toint }}</td>
  <td class="{{ 'g' if u.ticket_medio>=199 else 'r' }}">{{ u.ticket_medio|brl0 }}</td>
  <td class="{{ 'g' if u.fin_atual>=0.94 else 'r' }}">{{ u.fin_atual|pct }}</td>
  <td>{{ u.fin_30|pct }}</td><td>{{ u.fin_60|pct }}</td>
  <td class="{{ 'g' if u.frequencia>=0.75 else 'r' }}">{{ u.frequencia|pct }}</td>
  <td class="{{ 'g' if u.retencao>=0.94 else 'r' }}">{{ u.retencao|pct }}</td>
  <td class="sc">{{ u.rank_score }}</td>
  <td class="{{ 'g' if u.status=='bom' else 'r' }}">{{ 'BOM' if u.status=='bom' else 'RUIM' }}</td>
</tr>
{% endfor %}
</table>
{% elif modo=='comercial' %}
<table>
<tr><th>#</th><th style="text-align:left">Unidade</th><th style="text-align:left">Faturamento Comercial</th><th>Ticket Médio</th><th>Matrículas</th></tr>
{% for u in ranking %}
<tr>
  <td class="pos">{% if u.pos_crit==1 %}🥇{% elif u.pos_crit==2 %}🥈{% elif u.pos_crit==3 %}🥉{% else %}{{ u.pos_crit }}{% endif %}</td>
  <td class="nm">{{ u.nome }}</td>
  <td class="sc" style="text-align:left">{{ u.fat_comercial|brl }}</td>
  <td>{{ u.ticket_medio|brl0 }}</td>
  <td>{{ u.matriculas|toint }}</td>
</tr>
{% endfor %}
</table>
{% else %}
<table>
<tr><th>#</th><th style="text-align:left">Unidade</th><th>TPCv1</th><th>Canc. SCPC</th><th>Faturamento Comercial</th></tr>
{% for u in ranking %}
<tr>
  <td class="pos">{% if u.pos_crit==1 %}🥇{% elif u.pos_crit==2 %}🥈{% elif u.pos_crit==3 %}🥉{% else %}{{ u.pos_crit }}{% endif %}</td>
  <td class="nm">{{ u.nome }}</td>
  <td class="sc">{{ u.tpcv1|pct }}</td>
  <td class="{{ 'r' if u.canc_scpc>0 else 'g' }}">{{ u.canc_scpc|toint }}</td>
  <td>{{ u.fat_comercial|brl0 }}</td>
</tr>
{% endfor %}
</table>
<p style="text-align:center;color:#646c8c;font-size:.72rem;margin:4px 18px 0;font-weight:600">📉 No TPCv1, quanto menor melhor — empate é decidido pelo maior Faturamento Comercial.</p>
{% endif %}
<div class="rod">Gerado pelo sistema IFP Dashboard &nbsp;|&nbsp; {{ periodo }}</div>
{% if auto_print %}<script>window.addEventListener('load',function(){setTimeout(function(){window.print()},800)});</script>{% endif %}
</body></html>
"""

# ═════════════════════════════════════════════════════════════
# ROTAS
# ═════════════════════════════════════════════════════════════
def _calc_totais(unidades):
    tm=sum(u["matriculas"] for u in unidades)
    tfc=sum(u["fat_comercial"] for u in unidades)
    tft=sum(u["fat_total"] for u in unidades)
    ta=sum(u["ativos"] for u in unidades)
    rl=[u["retencao"] for u in unidades if u["retencao"]>0]
    fl=[u["frequencia"] for u in unidades if u["frequencia"]>0]
    fil=[u["fin_atual"] for u in unidades if u["fin_atual"]>0]
    tk=[u["ticket_medio"] for u in unidades if u["ticket_medio"]>0]
    return {"matriculas":tm,"ativos":ta,"fat_comercial":tfc,"fat_total":tft,
            "ticket_str":fmt_brl0(sum(tk)/len(tk) if tk else 0),
            "retencao_str":f"{(sum(rl)/len(rl) if rl else 0)*100:.1f}%",
            "freq_str":f"{(sum(fl)/len(fl) if fl else 0)*100:.1f}%",
            "fin_atual_str":f"{(sum(fil)/len(fil) if fil else 0)*100:.1f}%"}

@app.route("/fechamento", methods=["GET"])
def index():
    data,periodo,filename=get_cached_data()
    unidades=data or []
    totais=_calc_totais(unidades) if unidades else {}
    return render_template_string(TEMPLATE.replace("__CSS__",CSS),unidades=unidades,totais=totais,
        periodo=periodo,msg=request.args.get("msg",""),msg_ok=request.args.get("ok","1")=="1")

@app.route("/unidade/<int:idx>", methods=["GET"])
def unidade(idx):
    data,periodo,_=get_cached_data()
    if not data or idx>=len(data):
        return redirect("/fechamento?msg=Unidade+não+encontrada&ok=0")
    u=data[idx]
    # posição no ranking
    rk=ranquear(list(data))
    pos=next((x["posicao"] for x in rk if x["nome"]==u["nome"]),"—")
    return render_template_string(UNIDADE_TEMPLATE.replace("__CSS__",CSS),
        u=u,idx=idx,posicao=pos,periodo=periodo)

def _preparar_metricas(rk, modo):
    """Define metric_str e metric_sub exibidos no pódio para cada modo."""
    for u in rk:
        if modo=="geral":
            u["metric_str"]=f'{u["rank_score"]}'
            u["metric_sub"]=f'{u["score"]}/5 metas'
        elif modo=="comercial":
            u["metric_str"]=fmt_brl0(u.get("fat_comercial",0))
            u["metric_sub"]=f'{int(u.get("matriculas",0))} matrículas'
        elif modo=="tpcv1":
            u["metric_str"]=fmt_pct(u.get("tpcv1",0))
            u["metric_sub"]="menor é melhor"
    return rk

@app.route("/ranking", methods=["GET"])
def ranking():
    data,periodo,_=get_cached_data()
    rk=[]
    if data:
        for i,u in enumerate(data): u["idx_orig"]=i
        rk=_preparar_metricas(ranquear(list(data)),"geral")
    return render_template_string(RANKING_TEMPLATE.replace("__CSS__",CSS),
        ranking=rk,periodo=periodo,modo="geral",
        titulo="Ranking Geral de Unidades",
        subtitulo="🏆 Classificação geral por desempenho (5 indicadores vs metas)",
        pdf_url="/pdf/ranking")

@app.route("/ranking/comercial", methods=["GET"])
def ranking_comercial():
    data,periodo,_=get_cached_data()
    rk=[]
    if data:
        for i,u in enumerate(data): u["idx_orig"]=i
        rk=_preparar_metricas(ranquear_por(list(data),"fat_comercial",reverse=True),"comercial")
    return render_template_string(RANKING_TEMPLATE.replace("__CSS__",CSS),
        ranking=rk,periodo=periodo,modo="comercial",
        titulo="Ranking Comercial",
        subtitulo="💰 Classificação por Faturamento Comercial (maior → menor) — fonte VF",
        pdf_url="/pdf/ranking/comercial")

@app.route("/ranking/tpcv1", methods=["GET"])
def ranking_tpcv1():
    data,periodo,_=get_cached_data()
    rk=[]
    if data:
        for i,u in enumerate(data): u["idx_orig"]=i
        rk=_preparar_metricas(ranquear_por(list(data),"tpcv1",reverse=False,desempate="fat_comercial",desempate_reverse=True),"tpcv1")
    return render_template_string(RANKING_TEMPLATE.replace("__CSS__",CSS),
        ranking=rk,periodo=periodo,modo="tpcv1",
        titulo="Ranking TPCv1",
        subtitulo="📉 Classificação por TPCv1 (menor → melhor) — fonte VF",
        pdf_url="/pdf/ranking/tpcv1")

@app.route("/upload", methods=["POST"])
def upload():
    f=request.files.get("planilha")
    vf=request.files.get("vf")
    if not f or not f.filename:
        return redirect("/fechamento?msg=Envie+a+planilha+Excel&ok=0")
    if not vf or not vf.filename:
        return redirect("/fechamento?msg=O+arquivo+VF+(PDF)+é+obrigatório&ok=0")
    if not f.filename.lower().endswith((".xlsx",".xlsm")):
        return redirect("/fechamento?msg=Excel+inválido+(use+.xlsx)&ok=0")
    if not vf.filename.lower().endswith(".pdf"):
        return redirect("/fechamento?msg=VF+deve+ser+.pdf&ok=0")
    try:
        eb=f.read(); vb=vf.read()
        unidades=parse_tudo(eb,vb)
        m=re.search(r'(jan|fev|mar|abr|mai|jun|jul|ago|set|out|nov|dez)\w*[_\s]*\d{4}',vf.filename.lower())
        periodo=m.group(0).replace("_"," ").title() if m else vf.filename
        set_cached_data(unidades,periodo,f.filename,wb_bytes=eb)
        n_vf=sum(1 for u in unidades if u["fonte_vf"])
        return redirect(f"/fechamento?msg=✅+{len(unidades)}+unidades+({n_vf}+casadas+com+VF)&ok=1")
    except Exception as e:
        traceback.print_exc()
        return redirect(f"/fechamento?msg=Erro:+{str(e)[:130]}&ok=0")

@app.route("/pdf/todas")
def pdf_todas():
    data,periodo,_=get_cached_data()
    if not data: return redirect("/fechamento?msg=Sem+dados&ok=0")
    return Response(render_template_string(PDF_TEMPLATE,unidades=data,
        titulo="Relatório Geral — Todas as Unidades",periodo=periodo or "—",auto_print=False),mimetype="text/html")

@app.route("/pdf/unidade/<int:idx>")
def pdf_unidade(idx):
    data,periodo,_=get_cached_data()
    if not data or idx>=len(data): return redirect("/fechamento?msg=Não+encontrado&ok=0")
    u=data[idx]
    return Response(render_template_string(PDF_TEMPLATE,unidades=[u],
        titulo=f"Relatório — {u['nome']}",periodo=periodo or "—",auto_print=True),mimetype="text/html")

@app.route("/pdf/ranking")
def pdf_ranking():
    data,periodo,_=get_cached_data()
    if not data: return redirect("/fechamento?msg=Sem+dados&ok=0")
    rk=_preparar_metricas(ranquear(list(data)),"geral")
    return Response(render_template_string(PDF_RANKING,ranking=rk,periodo=periodo or "—",
        auto_print=True,modo="geral",titulo="Ranking Geral de Unidades"),mimetype="text/html")

@app.route("/pdf/ranking/comercial")
def pdf_ranking_comercial():
    data,periodo,_=get_cached_data()
    if not data: return redirect("/fechamento?msg=Sem+dados&ok=0")
    rk=_preparar_metricas(ranquear_por(list(data),"fat_comercial",reverse=True),"comercial")
    return Response(render_template_string(PDF_RANKING,ranking=rk,periodo=periodo or "—",
        auto_print=True,modo="comercial",titulo="Ranking Comercial"),mimetype="text/html")

@app.route("/pdf/ranking/tpcv1")
def pdf_ranking_tpcv1():
    data,periodo,_=get_cached_data()
    if not data: return redirect("/fechamento?msg=Sem+dados&ok=0")
    rk=_preparar_metricas(ranquear_por(list(data),"tpcv1",reverse=False,desempate="fat_comercial",desempate_reverse=True),"tpcv1")
    return Response(render_template_string(PDF_RANKING,ranking=rk,periodo=periodo or "—",
        auto_print=True,modo="tpcv1",titulo="Ranking TPCv1 (menor → melhor)"),mimetype="text/html")

@app.route("/api/dados")
def api_dados():
    data,periodo,filename=get_cached_data()
    return jsonify({"periodo":periodo,"filename":filename,"total":len(data or []),"unidades":data or []})

# ═════════════════════════════════════════════════════════════
# HOME — escolha entre Fechamento Mensal e Período Decendial
# ═════════════════════════════════════════════════════════════
HOME_TEMPLATE = r"""
<!DOCTYPE html><html lang="pt-br"><head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>IFP – Plataforma</title><style>__CSS__
.home-wrap{max-width:980px;margin:0 auto;padding:60px 26px}
.home-hero{text-align:center;margin-bottom:46px}
.home-hero .big-logo{width:92px;height:92px;background:#fff;border-radius:24px;display:inline-flex;align-items:center;justify-content:center;box-shadow:var(--shadow-lg);margin-bottom:20px}
.home-hero .big-logo span{font-family:'Sora';font-size:2rem;font-weight:800;color:var(--red);letter-spacing:-2px}
.home-hero h1{font-size:2.1rem;font-weight:800;color:var(--blue);letter-spacing:-.6px}
.home-hero p{font-size:1rem;color:var(--txt2);margin-top:8px;font-weight:500}
.home-grid{display:grid;grid-template-columns:1fr 1fr;gap:24px}
.home-card{border-radius:24px;padding:38px 32px;color:#fff;text-decoration:none;position:relative;overflow:hidden;box-shadow:var(--shadow-lg);transition:.22s;display:flex;flex-direction:column;min-height:260px}
.home-card:hover{transform:translateY(-6px)}
.home-card::after{content:"";position:absolute;inset:0;background:radial-gradient(500px 200px at 90% 130%,rgba(255,255,255,.16),transparent)}
.home-card.mensal{background:linear-gradient(150deg,var(--red-dk),var(--red) 70%,var(--red-br))}
.home-card.dec{background:linear-gradient(150deg,var(--blue),var(--blue-2) 70%,var(--blue-lt))}
.home-card .hc-ico{font-size:3rem;margin-bottom:16px;position:relative}
.home-card h2{font-size:1.5rem;font-weight:800;position:relative;letter-spacing:-.4px}
.home-card p{font-size:.92rem;opacity:.9;margin-top:10px;line-height:1.5;position:relative;flex:1}
.home-card .hc-cta{margin-top:18px;font-weight:800;font-size:.92rem;display:inline-flex;align-items:center;gap:8px;position:relative;background:rgba(255,255,255,.2);border:1.5px solid rgba(255,255,255,.4);border-radius:30px;padding:9px 20px;align-self:flex-start}
@media(max-width:680px){.home-grid{grid-template-columns:1fr}}
</style></head><body>
<div class="hdr"><div class="hdr-in">
  <div class="logo-wrap"><div class="logo-img"><span>IFP</span></div>
  <div class="logo-txt"><h1>Plataforma de Gestão</h1><p>Instituto de Formação Profissional</p></div></div>
</div></div>
<div class="home-wrap">
  <div class="home-hero">
    <div class="big-logo"><span>IFP</span></div>
    <h1>Bem-vindo à Plataforma IFP</h1>
    <p>Escolha o módulo que deseja acessar</p>
  </div>
  <div class="home-grid">
    <a href="/fechamento" class="home-card mensal">
      <span class="hc-ico">📊</span>
      <h2>Fechamento Mensal</h2>
      <p>Painel completo das unidades com indicadores, rankings (geral, comercial e TPCv1) e relatórios em PDF. Compara Excel + VF do mês fechado.</p>
      <span class="hc-cta">Acessar módulo ➜</span>
    </a>
    <a href="/decendial" class="home-card dec">
      <span class="hc-ico">🗓️</span>
      <h2>Período Decendial</h2>
      <p>Cadastre as metas mensais de cada unidade e acompanhe o andamento a cada 10 dias, exportando o VF 3× no mês para ver o quão perto está de bater as metas.</p>
      <span class="hc-cta">Acessar módulo ➜</span>
    </a>
  </div>
</div>
</body></html>
"""

@app.route("/", methods=["GET"])
def home():
    return render_template_string(HOME_TEMPLATE.replace("__CSS__",CSS))

# ═════════════════════════════════════════════════════════════
# DECENDIAL — CSS extra + navegação
# ═════════════════════════════════════════════════════════════
CSS_DEC = r"""
.dec-wrap{max-width:1280px;margin:24px auto 60px;padding:0 26px}
.dec-nav{background:var(--blue);padding:0 26px;box-shadow:0 3px 14px rgba(22,32,94,.2)}
.dec-nav-in{max-width:1280px;margin:0 auto;display:flex;align-items:center;gap:4px;flex-wrap:wrap}
.dec-nav a{padding:14px 18px;color:rgba(255,255,255,.62);text-decoration:none;font-size:.86rem;font-weight:700;border-bottom:3px solid transparent;margin-bottom:-1px;transition:.18s}
.dec-nav a:hover{color:#fff;background:rgba(255,255,255,.06)}
.dec-nav a.ativo{color:#fff;border-bottom-color:var(--red-br)}
.dec-toolbar{display:flex;align-items:center;gap:12px;flex-wrap:wrap;margin-bottom:20px}
.dec-unidade-sel{padding:9px 14px;border-radius:11px;border:1.5px solid var(--borda);background:#fff;font-size:.86rem;font-weight:700;color:var(--blue);font-family:inherit;min-width:240px}
.dec-card{background:#fff;border-radius:var(--r);box-shadow:var(--shadow);border:1px solid var(--borda);overflow:hidden;margin-bottom:20px}
.dec-card-h{background:linear-gradient(135deg,var(--blue),var(--blue-2));color:#fff;padding:16px 22px;display:flex;align-items:center;justify-content:space-between;gap:12px;flex-wrap:wrap}
.dec-card-h h3{font-size:1.05rem;font-weight:800}
.dec-card-h .sub{font-size:.76rem;opacity:.85;font-weight:500;margin-top:2px}
.metas-table{width:100%;border-collapse:collapse;font-size:.82rem}
.metas-table th{background:var(--blue-soft);color:var(--blue);font-size:.66rem;font-weight:800;text-transform:uppercase;letter-spacing:.04em;padding:11px 10px;text-align:center;border-bottom:2px solid var(--borda)}
.metas-table th:first-child{text-align:left;padding-left:22px}
.metas-table td{padding:7px 10px;border-bottom:1px solid var(--borda);text-align:center}
.metas-table td.lbl{text-align:left;padding-left:22px;font-weight:700;color:var(--txt);font-size:.8rem}
.metas-table tr:nth-child(even){background:var(--bg-2)}
.metas-table tr:hover{background:var(--blue-soft)}
.metas-table input{width:100%;max-width:130px;padding:7px 9px;border:1.5px solid var(--borda);border-radius:8px;font-size:.82rem;font-family:inherit;text-align:right;background:#fff;color:var(--txt);transition:.14s}
.metas-table input:focus{border-color:var(--red);outline:none;box-shadow:0 0 0 3px var(--red-lt)}
.metas-table .pct-cell{font-weight:800;color:var(--blue);font-family:'Sora'}
.metas-table .pct-cell.g{color:var(--verde)}.metas-table .pct-cell.r{color:var(--rose)}
.dec-total{display:flex;align-items:center;justify-content:space-between;padding:16px 22px;background:linear-gradient(135deg,var(--blue-soft),#dfe3f7);border-top:2px solid var(--borda)}
.dec-total-l{font-size:.95rem;font-weight:800;color:var(--blue);text-transform:uppercase;letter-spacing:.03em}
.dec-total-v{font-size:1.5rem;font-weight:800;color:var(--blue);font-family:'Sora'}
.dec-actions{display:flex;gap:12px;flex-wrap:wrap;margin-top:18px}
.btn-amber{background:var(--amber);color:#fff}.btn-amber:hover{background:#a85308}
.btn-green{background:var(--verde);color:#fff}.btn-green:hover{background:#15602f}
.btn-clear{background:#fff;color:var(--rose);border:1.5px solid var(--rose)}.btn-clear:hover{background:var(--rose-bg)}
.dec-export-box{background:linear-gradient(150deg,var(--red-dk),var(--red) 70%,var(--red-br));border-radius:var(--r);padding:30px 28px;color:#fff;text-align:center;margin-bottom:24px;box-shadow:var(--shadow-red);position:relative;overflow:hidden}
.dec-export-box::after{content:"";position:absolute;inset:0;background:radial-gradient(500px 200px at 85% 130%,rgba(255,255,255,.14),transparent)}
.dec-export-box h3{font-size:1.4rem;font-weight:800;position:relative}
.dec-export-box p{font-size:.9rem;opacity:.92;margin-top:8px;position:relative}
.dec-export-form{margin-top:18px;display:flex;align-items:center;justify-content:center;gap:12px;flex-wrap:wrap;position:relative}
.dec-export-form select,.dec-export-form input[type=file]{padding:11px 14px;border-radius:11px;border:none;font-size:.86rem;font-weight:700;font-family:inherit}
.dec-export-form input[type=file]{background:rgba(255,255,255,.16);color:#fff;border:1.5px solid rgba(255,255,255,.4)}
.dec-export-form input[type=file]::-webkit-file-upload-button{background:#fff;color:var(--red);border:none;border-radius:8px;padding:6px 12px;font-weight:700;cursor:pointer;margin-right:8px}
.btn-export-big{background:#fff;color:var(--red);font-size:1rem;font-weight:800;padding:13px 30px;border-radius:13px;border:none;cursor:pointer;box-shadow:0 6px 20px rgba(0,0,0,.2)}
.btn-export-big:hover{transform:translateY(-2px)}
.dec-units-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(330px,1fr));gap:16px}
.dec-unit{background:#fff;border-radius:var(--r);box-shadow:var(--shadow);border:1px solid var(--borda);overflow:hidden}
.dec-unit-h{padding:14px 18px;background:linear-gradient(135deg,var(--blue),var(--blue-2));color:#fff;font-weight:800;font-size:.95rem}
.dec-unit-b{padding:14px 16px}
.dec-slot{display:flex;align-items:center;justify-content:space-between;gap:10px;padding:9px 12px;border-radius:10px;margin-bottom:8px;border:1.5px solid var(--borda);font-size:.82rem;font-weight:700}
.dec-slot.feito{background:var(--verde-bg);border-color:var(--verde-lt)}
.dec-slot.pendente{background:var(--bg-2)}
.dec-slot .slot-lbl{color:var(--txt);display:flex;align-items:center;gap:8px}
.dec-slot .slot-st{font-size:.7rem;font-weight:800;padding:3px 10px;border-radius:30px}
.dec-slot.feito .slot-st{background:var(--verde);color:#fff}
.dec-slot.pendente .slot-st{background:#dfe3ee;color:var(--txt2)}
.dec-slot a{text-decoration:none;color:var(--blue);font-size:.72rem;font-weight:800}
.dec-empty{text-align:center;padding:50px 20px;color:var(--txt2)}
.dec-empty .ico{font-size:3rem;margin-bottom:12px}
.res-table{width:100%;border-collapse:collapse;font-size:.82rem;margin-top:6px}
.res-table th{background:var(--blue);color:#fff;font-size:.64rem;font-weight:800;text-transform:uppercase;padding:9px 8px}
.res-table th:first-child{text-align:left;padding-left:16px}
.res-table td{padding:8px;border-bottom:1px solid var(--borda);text-align:center}
.res-table td.lbl{text-align:left;padding-left:16px;font-weight:700}
.res-table .ating{font-weight:800;font-family:'Sora'}
.res-table .ating.g{color:var(--verde)}.res-table .ating.am{color:var(--amber)}.res-table .ating.r{color:var(--rose)}
.dec-export-box .dec-progress{display:inline-block;margin-top:14px;position:relative;background:rgba(255,255,255,.18);border:1.5px solid rgba(255,255,255,.4);border-radius:30px;padding:7px 18px;font-size:.82rem;font-weight:800}
.dec-banner{display:flex;align-items:flex-start;gap:12px;padding:15px 20px;border-radius:14px;margin-bottom:18px;font-size:.9rem;font-weight:700;box-shadow:var(--shadow);line-height:1.45}
.dec-banner .db-ico{font-size:1.3rem;flex-shrink:0;line-height:1.2}
.dec-banner.ok{background:var(--verde-bg);color:var(--verde);border:1.5px solid var(--verde-lt)}
.dec-banner.err{background:linear-gradient(135deg,var(--red),var(--red-dk));color:#fff;border:1.5px solid var(--red-dk)}
"""

def dec_nav(ativo):
    return f"""<div class="dec-nav"><div class="dec-nav-in">
      <a href="/">🏠 Início</a>
      <a href="/decendial" class="{'ativo' if ativo=='metas' else ''}">🎯 Metas das Unidades</a>
      <a href="/decendial/acompanhamento" class="{'ativo' if ativo=='acomp' else ''}">🗓️ Acompanhamento (10 em 10 dias)</a>
    </div></div>"""

# ═════════════════════════════════════════════════════════════
# DECENDIAL — formulário de metas
# ═════════════════════════════════════════════════════════════
DEC_METAS_TEMPLATE = r"""
<!DOCTYPE html><html lang="pt-br"><head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>IFP – Metas Decendiais</title><style>__CSS__</style></head><body>
<div class="hdr"><div class="hdr-in">
  <div class="logo-wrap"><div class="logo-img"><span>IFP</span></div>
  <div class="logo-txt"><h1>Período Decendial — Metas</h1><p>Instituto de Formação Profissional</p></div></div>
</div></div>
__NAV__
<div class="dec-wrap">
{% if not unidades %}
  <div class="dec-card"><div class="dec-empty"><div class="ico">📊</div>
    <h3 style="color:var(--blue);font-size:1.1rem;margin-bottom:8px">Nenhuma unidade carregada ainda</h3>
    <p>Para cadastrar as metas, primeiro carregue a planilha + VF no módulo <a href="/fechamento" style="color:var(--red);font-weight:800">Fechamento Mensal</a>. As unidades de lá aparecerão aqui automaticamente.</p>
  </div></div>
{% else %}
  <div class="dec-toolbar">
    <span class="flabel">Selecione a unidade:</span>
    <select class="dec-unidade-sel" onchange="window.location='/decendial?u='+encodeURIComponent(this.value)">
      {% for nome in unidades %}
      <option value="{{ nome }}" {{ 'selected' if nome==unidade_sel else '' }}>{{ nome }}</option>
      {% endfor %}
    </select>
    {% if msg %}<span class="smsg {{ 'ok' if msg_ok else 'err' }}">{{ msg }}</span>{% endif %}
    <div style="flex:1"></div>
    <a href="/decendial/pdf?u={{ unidade_sel|urlencode }}" target="_blank" class="btn btn-blue">⬇ PDF</a>
    <a href="/decendial/excel?u={{ unidade_sel|urlencode }}" class="btn btn-green">⬇ Excel</a>
  </div>
  <form method="POST" action="/decendial/salvar">
  <input type="hidden" name="unidade" value="{{ unidade_sel }}">
  <div class="dec-card">
    <div class="dec-card-h">
      <div><h3>🎯 {{ unidade_sel }}</h3><div class="sub">Metas do mês — preencha uma vez; ficam salvas para o acompanhamento decendial</div></div>
    </div>
    <table class="metas-table">
      <thead><tr>
        <th>% Atingido</th><th>Mês Anterior</th><th>Meta Mês</th><th>Quantidade</th><th>Valor</th>
      </tr></thead>
      <tbody>
      {% for l in linhas %}
        <tr>
          <td class="lbl">{{ l.label }}</td>
          <td><input type="text" name="{{ l.id }}__mes_anterior" value="{{ metas.get(l.id,{}).get('mes_anterior','') }}" placeholder="0" inputmode="decimal"></td>
          <td><input type="text" name="{{ l.id }}__meta_mes" value="{{ metas.get(l.id,{}).get('meta_mes','') }}" placeholder="0" inputmode="decimal"></td>
          <td><input type="text" name="{{ l.id }}__quantidade" value="{{ metas.get(l.id,{}).get('quantidade','') }}" placeholder="0" inputmode="decimal"></td>
          <td><input type="text" name="{{ l.id }}__valor" value="{{ metas.get(l.id,{}).get('valor','') }}" placeholder="0,00" inputmode="decimal"></td>
        </tr>
      {% endfor %}
      </tbody>
    </table>
    <div class="dec-total">
      <span class="dec-total-l">Total Recebido</span>
      <span class="dec-total-v">{{ total_receb|brl }}</span>
    </div>
  </div>
  <div class="dec-actions">
    <button type="submit" class="btn btn-red">💾 Salvar metas desta unidade</button>
    <a href="/decendial/limpar?u={{ unidade_sel|urlencode }}" class="btn btn-clear"
       onclick="return confirm('Limpar TODOS os campos desta unidade?')">🗑️ Limpar campos</a>
    <a href="/decendial/limpar_tudo" class="btn btn-clear"
       onclick="return confirm('Limpar os campos de TODAS as unidades? Esta ação não pode ser desfeita.')">🗑️ Limpar tudo (todas unidades)</a>
  </div>
  </form>
{% endif %}
</div>
</body></html>
"""

@app.route("/decendial", methods=["GET"])
def decendial():
    nomes = dec_unidades()
    unidade_sel = request.args.get("u") or (nomes[0] if nomes else "")
    metas = metas_da_unidade(unidade_sel)
    html = DEC_METAS_TEMPLATE.replace("__CSS__",CSS+CSS_DEC).replace("__NAV__",dec_nav("metas"))
    return render_template_string(html, unidades=nomes, unidade_sel=unidade_sel,
        linhas=LINHAS_DEC, metas=metas, total_receb=total_recebido(unidade_sel),
        msg=request.args.get("msg",""), msg_ok=request.args.get("ok","1")=="1")

@app.route("/decendial/salvar", methods=["POST"])
def decendial_salvar():
    nome = request.form.get("unidade","").strip()
    if not nome: return redirect("/decendial?msg=Selecione+uma+unidade&ok=0")
    bloco = {}
    for l in LINHAS_DEC:
        d={}
        for c in CAMPOS_DEC:
            v = request.form.get(f"{l['id']}__{c}","").strip()
            if v: d[c]=v
        if d: bloco[l["id"]]=d
    _dec["metas"][nome]=bloco
    # recalcula o score do fechamento com as novas metas (se já houver dados)
    data,per,fn = get_cached_data()
    if data:
        for u in data:
            if u["nome"]==nome: calcular_score(u)
    return redirect(f"/decendial?u={nome}&msg=✅+Metas+salvas&ok=1")

@app.route("/decendial/limpar", methods=["GET"])
def decendial_limpar():
    nome = request.args.get("u","").strip()
    if nome in _dec["metas"]: _dec["metas"].pop(nome)
    data,_,_=get_cached_data()
    if data:
        for u in data:
            if u["nome"]==nome: calcular_score(u)
    return redirect(f"/decendial?u={nome}&msg=Campos+limpos&ok=1")

@app.route("/decendial/limpar_tudo", methods=["GET"])
def decendial_limpar_tudo():
    _dec["metas"].clear()
    data,_,_=get_cached_data()
    if data:
        for u in data: calcular_score(u)
    return redirect("/decendial?msg=Todos+os+campos+foram+limpos&ok=1")

# ═════════════════════════════════════════════════════════════
# DECENDIAL — % atingido para uma linha (Valor / Meta-em-Valor)
# ═════════════════════════════════════════════════════════════
def pct_atingido_linha(meta_linha):
    """% atingido = quantidade/valor realizado vs meta. Aqui usamos Valor/MetaMês
       quando ambos existem; senão Quantidade/MetaMês."""
    meta = to_float(meta_linha.get("meta_mes",0))
    val  = to_float(meta_linha.get("valor",0))
    qtd  = to_float(meta_linha.get("quantidade",0))
    base = val if val>0 else qtd
    if meta>0: return base/meta
    return 0.0

# ═════════════════════════════════════════════════════════════
# DECENDIAL — PDF e Excel das metas preenchidas
# ═════════════════════════════════════════════════════════════
DEC_PDF_METAS = r"""
<!DOCTYPE html><html lang="pt-br"><head><meta charset="UTF-8"><title>IFP – Metas {{ unidade }}</title>
<link href="https://fonts.googleapis.com/css2?family=Sora:wght@600;700;800&family=Plus+Jakarta+Sans:wght@400;600;700;800&display=swap" rel="stylesheet">
<style>
*{box-sizing:border-box;margin:0;padding:0}body{font-family:'Plus Jakarta Sans',sans-serif;color:#141a36;font-size:11px}
.ph{background:linear-gradient(120deg,#16205e,#1f2d7a);color:#fff;padding:18px 24px;display:flex;align-items:center;justify-content:space-between;margin-bottom:18px}
.ph-logo{width:42px;height:42px;background:#fff;border-radius:10px;display:flex;align-items:center;justify-content:center;font-family:'Sora';font-weight:800;color:#c0021c}
.ph-r h2{font-size:1rem;font-weight:800;text-align:right}.ph-r p{font-size:.68rem;opacity:.85;text-align:right;margin-top:2px}
.no-print{text-align:center;margin-bottom:14px}.no-print button{border:none;border-radius:10px;padding:10px 26px;font-size:.88rem;font-weight:700;cursor:pointer;margin:0 4px}
.btn-p{background:#c0021c;color:#fff}.btn-c{background:#16205e;color:#fff}
table{border-collapse:collapse;width:calc(100% - 36px);margin:0 18px 12px;font-size:11px}
th,td{border:1px solid #e2e6f2;padding:8px 10px;text-align:right}
th{background:#16205e;color:#fff;font-size:.64rem;text-transform:uppercase;text-align:center}
td.lbl{text-align:left;font-weight:700}
tr:nth-child(even){background:#f6f8fc}
.tot{display:flex;justify-content:space-between;margin:0 18px;padding:14px 18px;background:#eef0fb;border-radius:10px;font-weight:800;color:#16205e}
.tot .v{font-family:'Sora';font-size:1.3rem}
.rod{text-align:center;color:#9aa3b2;font-size:.64rem;padding:14px 0;margin:0 18px}
@media print{body{-webkit-print-color-adjust:exact;print-color-adjust:exact}.no-print{display:none}}
</style></head><body>
<div class="ph"><div class="ph-logo">IFP</div><div class="ph-r"><h2>Metas do Mês — {{ unidade }}</h2><p>Período Decendial &nbsp;|&nbsp; Instituto de Formação Profissional</p></div></div>
<div class="no-print"><button class="btn-p" onclick="window.print()">🖨️ Imprimir / Salvar PDF</button><button class="btn-c" onclick="window.close()">✕ Fechar</button></div>
<table>
<tr><th style="text-align:left">% Atingido</th><th>Mês Anterior</th><th>Meta Mês</th><th>Quantidade</th><th>Valor</th></tr>
{% for l in linhas %}
<tr>
  <td class="lbl">{{ l.label }}</td>
  <td>{{ metas.get(l.id,{}).get('mes_anterior','—') }}</td>
  <td>{{ metas.get(l.id,{}).get('meta_mes','—') }}</td>
  <td>{{ metas.get(l.id,{}).get('quantidade','—') }}</td>
  <td>{{ metas.get(l.id,{}).get('valor','—') }}</td>
</tr>
{% endfor %}
</table>
<div class="tot"><span>TOTAL RECEBIDO</span><span class="v">{{ total_receb|brl }}</span></div>
<div class="rod">Gerado pelo sistema IFP Dashboard — Período Decendial</div>
<script>window.addEventListener('load',function(){setTimeout(function(){window.print()},700)});</script>
</body></html>
"""

@app.route("/decendial/pdf", methods=["GET"])
def decendial_pdf():
    nome = request.args.get("u","").strip()
    if not nome: return redirect("/decendial?msg=Selecione+uma+unidade&ok=0")
    html = DEC_PDF_METAS.replace("__CSS__","")
    return Response(render_template_string(html, unidade=nome, linhas=LINHAS_DEC,
        metas=metas_da_unidade(nome), total_receb=total_recebido(nome)), mimetype="text/html")

@app.route("/decendial/excel", methods=["GET"])
def decendial_excel():
    nome = request.args.get("u","").strip()
    if not nome: return redirect("/decendial?msg=Selecione+uma+unidade&ok=0")
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    wb = Workbook(); ws = wb.active; ws.title = "Metas"[:31]
    azul = "FF16205E"; az_claro="FFEEF0FB"
    hdr_font=Font(bold=True,color="FFFFFFFF",size=10)
    hdr_fill=PatternFill("solid",fgColor=azul)
    bd=Border(*[Side(style="thin",color="FFE2E6F2")]*4)
    cab=["% Atingido","Mês Anterior","Meta Mês","Quantidade","Valor"]
    ws.append([f"Metas do Mês — {nome}"])
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=5)
    ws.cell(1,1).font=Font(bold=True,size=13,color=azul)
    ws.append([])
    ws.append(cab)
    for j in range(1,6):
        c=ws.cell(ws.max_row,j); c.font=hdr_font; c.fill=hdr_fill
        c.alignment=Alignment(horizontal="center"); c.border=bd
    m=metas_da_unidade(nome)
    for l in LINHAS_DEC:
        d=m.get(l["id"],{})
        ws.append([l["label"], d.get("mes_anterior",""), d.get("meta_mes",""),
                   d.get("quantidade",""), d.get("valor","")])
        for j in range(1,6):
            cc=ws.cell(ws.max_row,j); cc.border=bd
            if j==1: cc.font=Font(bold=True,size=9)
            else: cc.alignment=Alignment(horizontal="right")
    ws.append([])
    ws.append(["TOTAL RECEBIDO","","","", total_recebido(nome)])
    r=ws.max_row
    ws.cell(r,1).font=Font(bold=True,color=azul); ws.cell(r,1).fill=PatternFill("solid",fgColor=az_claro)
    ws.cell(r,5).font=Font(bold=True,color=azul); ws.cell(r,5).fill=PatternFill("solid",fgColor=az_claro)
    ws.cell(r,5).number_format='#,##0.00'
    ws.column_dimensions["A"].width=46
    for col in "BCDE": ws.column_dimensions[col].width=16
    bio=io.BytesIO(); wb.save(bio); bio.seek(0)
    safe="".join(ch for ch in nome if ch.isalnum() or ch in " -_").strip().replace(" ","_")
    return Response(bio.read(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition":f'attachment; filename="metas_{safe}.xlsx"'})

# ═════════════════════════════════════════════════════════════
# DECENDIAL — Acompanhamento (3 decêndios por unidade)
# ═════════════════════════════════════════════════════════════
DEC_ACOMP_TEMPLATE = r"""
<!DOCTYPE html><html lang="pt-br"><head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>IFP – Acompanhamento Decendial</title><style>__CSS__</style></head><body>
<div class="hdr"><div class="hdr-in">
  <div class="logo-wrap"><div class="logo-img"><span>IFP</span></div>
  <div class="logo-txt"><h1>Período Decendial — Acompanhamento</h1><p>Instituto de Formação Profissional</p></div></div>
</div></div>
__NAV__
<div class="dec-wrap">
{% if not unidades %}
  <div class="dec-card"><div class="dec-empty"><div class="ico">🗓️</div>
    <h3 style="color:var(--blue);font-size:1.1rem;margin-bottom:8px">Nenhuma unidade carregada</h3>
    <p>Carregue a planilha + VF no <a href="/fechamento" style="color:var(--red);font-weight:800">Fechamento Mensal</a> primeiro.</p>
  </div></div>
{% else %}
  {% if msg %}
  <div class="dec-banner {{ msg_tipo }}">
    <span class="db-ico">{% if msg_tipo=='err' %}⚠️{% else %}✅{% endif %}</span>
    <span>{{ msg }}</span>
  </div>
  {% endif %}
  {% if n_com_metas==0 %}
  <div class="dec-banner err">
    <span class="db-ico">📝</span>
    <span>Você ainda <strong>não cadastrou metas</strong> de nenhuma unidade. Vá em <a href="/decendial" style="color:#fff;text-decoration:underline;font-weight:800">Metas das Unidades</a>, preencha e salve. Só depois o "Exportar VF" terá com o que comparar.</span>
  </div>
  {% endif %}
  <div class="dec-export-box">
    <h3>📤 Exportar VF do período (10 dias)</h3>
    <p>Envie o VF correspondente a um decêndio. O sistema compara com as metas cadastradas e salva o resultado por unidade.<br>
       São <strong>3 envios no mês</strong>: dia 1 ao 10, dia 11 ao 21, e dia 22 ao último dia.</p>
    <div class="dec-progress">📋 {{ n_com_metas }} de {{ n_total }} unidades com metas cadastradas</div>
    <form class="dec-export-form" method="POST" action="/decendial/exportar_vf" enctype="multipart/form-data">
      <select name="decendio" required>
        {% for d in decendios %}<option value="{{ d.id }}">{{ d.label }}</option>{% endfor %}
      </select>
      <input type="file" name="vf" accept=".pdf,application/pdf" required>
      <button type="submit" class="btn-export-big">📊 Exportar VF e comparar</button>
    </form>
  </div>
  <div class="dec-units-grid">
    {% for nome in unidades %}
    {% set tem_metas = nome in resultados or false %}
    <div class="dec-unit">
      <div class="dec-unit-h">{{ nome }}</div>
      <div class="dec-unit-b">
        {% for d in decendios %}
          {% set feito = resultados.get(nome,{}).get(d.id) %}
          <div class="dec-slot {{ 'feito' if feito else 'pendente' }}">
            <span class="slot-lbl">{% if feito %}✅{% else %}⏳{% endif %} {{ d.label }}</span>
            {% if feito %}
              <a href="/decendial/resultado?u={{ nome|urlencode }}&d={{ d.id }}">ver resultado ➜</a>
            {% else %}
              <span class="slot-st">pendente</span>
            {% endif %}
          </div>
        {% endfor %}
      </div>
    </div>
    {% endfor %}
  </div>
{% endif %}
</div>
</body></html>
"""

@app.route("/decendial/acompanhamento", methods=["GET"])
def decendial_acomp():
    nomes = dec_unidades()
    com_metas = [nm for nm in nomes if _dec["metas"].get(nm)]
    html = DEC_ACOMP_TEMPLATE.replace("__CSS__",CSS+CSS_DEC).replace("__NAV__",dec_nav("acomp"))
    return render_template_string(html, unidades=nomes, decendios=DECENDIOS,
        resultados=_dec["resultados"], msg=request.args.get("msg",""),
        msg_tipo=request.args.get("t","ok"),
        n_com_metas=len(com_metas), n_total=len(nomes))

@app.route("/decendial/exportar_vf", methods=["POST"])
def decendial_exportar_vf():
    dec_id = request.form.get("decendio","").strip()
    vf = request.files.get("vf")
    if dec_id not in [d["id"] for d in DECENDIOS]:
        return redirect("/decendial/acompanhamento?t=err&msg=Selecione+um+período+válido")
    if not vf or not vf.filename:
        return redirect("/decendial/acompanhamento?t=err&msg=Nenhum+arquivo+escolhido.+Clique+em+'Escolher+arquivo'+e+selecione+o+VF+(PDF)")
    if not vf.filename.lower().endswith(".pdf"):
        return redirect("/decendial/acompanhamento?t=err&msg=O+arquivo+precisa+ser+um+VF+em+PDF")

    # Verifica se há alguma meta cadastrada antes de processar
    nomes = dec_unidades()
    com_metas = [nm for nm in nomes if _dec["metas"].get(nm)]
    if not com_metas:
        return redirect("/decendial/acompanhamento?t=err&msg=⚠️+Cadastre+as+metas+de+pelo+menos+uma+unidade+na+aba+'Metas+das+Unidades'+antes+de+exportar+o+VF")

    try:
        vf_dados = parse_vf_pdf(vf.read())
    except Exception as e:
        return redirect(f"/decendial/acompanhamento?t=err&msg=Erro+ao+ler+VF:+{str(e)[:80]}")
    if not vf_dados:
        return redirect("/decendial/acompanhamento?t=err&msg=Não+consegui+ler+dados+do+VF.+Confira+se+é+o+arquivo+correto")

    # Para cada unidade com metas, compara o VF (10 dias) com a meta do mês
    n=0; sem_match=[]
    for nome in com_metas:
        metas = _dec["metas"].get(nome)
        chave = norm_nome(nome)
        v = vf_dados.get(chave)
        if not v:
            sem_match.append(nome); continue
        # mapeia campos do VF para as linhas do formulário
        realizado = {
            "matricula":      v["fat_comercial"],   # valor comercial realizado
            "proj_comercial": v["fat_comercial"],
            "atual":          v["fin_atual"]*100,
            "d30":            v["fin_30"]*100,
            "d60":            v["fin_60"]*100,
            "cancelamentos":  v["cancelados"],
            "scpc":           v["canc_scpc"],
        }
        linhas_cmp=[]
        for l in LINHAS_DEC:
            meta_linha = metas.get(l["id"],{})
            meta_v = to_float(meta_linha.get("meta_mes",0))
            real = realizado.get(l["id"], 0.0)
            pct = (real/meta_v) if meta_v>0 else 0.0
            linhas_cmp.append({"id":l["id"],"label":l["label"],"tipo":l["tipo"],
                "meta":meta_v,"realizado":real,"pct":pct})
        _dec["resultados"].setdefault(nome,{})[dec_id]={
            "decendio":dec_id,
            "decendio_label":next(d["label"] for d in DECENDIOS if d["id"]==dec_id),
            "linhas":linhas_cmp,
            "fat_comercial":v["fat_comercial"],"fat_total":v["fat_total"],
            "matriculas":v["matriculas"],"fin_atual":v["fin_atual"],
            "fin_30":v["fin_30"],"fin_60":v["fin_60"],"tpcv1":v["tpcv1"],
            "cancelados":v["cancelados"],"canc_scpc":v["canc_scpc"],
        }
        n+=1
    lbl=next(d["label"] for d in DECENDIOS if d["id"]==dec_id)
    falt = len(nomes)-len(com_metas)
    msg = f"✅+{lbl}:+{n}+unidade(s)+comparada(s)+com+o+VF"
    if falt>0: msg += f".+{falt}+unidade(s)+ainda+sem+metas+cadastradas"
    if sem_match: msg += f".+{len(sem_match)}+com+metas+mas+não+encontrada(s)+no+VF"
    tipo = "ok" if n>0 else "err"
    if n==0:
        msg = "Nenhuma+unidade+foi+comparada:+as+unidades+com+metas+não+foram+encontradas+neste+VF"
    return redirect(f"/decendial/acompanhamento?t={tipo}&msg={msg}")

DEC_RESULT_TEMPLATE = r"""
<!DOCTYPE html><html lang="pt-br"><head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>IFP – Resultado {{ unidade }}</title><style>__CSS__</style></head><body>
<div class="hdr"><div class="hdr-in">
  <div class="logo-wrap"><div class="logo-img"><span>IFP</span></div>
  <div class="logo-txt"><h1>Resultado — {{ unidade }}</h1><p>{{ res.decendio_label }} · acompanhamento decendial</p></div></div>
</div></div>
__NAV__
<div class="dec-wrap">
  <a href="/decendial/acompanhamento" class="detail-back">⬅ Voltar ao acompanhamento</a>
  <div class="dec-card">
    <div class="dec-card-h">
      <div><h3>🗓️ {{ res.decendio_label }}</h3><div class="sub">Comparação do VF (período de 10 dias) com as metas do mês</div></div>
      <a href="/decendial/resultado/pdf?u={{ unidade|urlencode }}&d={{ res.decendio }}" target="_blank" class="btn btn-red">⬇ Baixar PDF</a>
    </div>
    <table class="res-table">
      <thead><tr><th>Indicador</th><th>Meta do Mês</th><th>Realizado (10d)</th><th>% Atingido</th><th>Progresso</th></tr></thead>
      <tbody>
      {% for l in res.linhas %}
        {% if l.meta>0 or l.realizado>0 %}
        {% set p = (l.pct*100) %}
        {% set cl = 'g' if p>=100 else ('am' if p>=60 else 'r') %}
        <tr>
          <td class="lbl">{{ l.label }}</td>
          <td>{% if l.tipo=='pct' %}{{ (l.meta)|pct if l.meta<=1.5 else '%.1f%%'|format(l.meta) }}{% else %}{{ l.meta|brl0 if l.meta>=100 else l.meta|toint }}{% endif %}</td>
          <td>{% if l.tipo=='pct' %}{{ '%.1f%%'|format(l.realizado) }}{% else %}{{ l.realizado|brl0 if l.realizado>=100 else l.realizado|toint }}{% endif %}</td>
          <td class="ating {{ cl }}">{{ '%.0f%%'|format(p) }}</td>
          <td><div class="minibar"><div style="width:{{ [p,100]|min }}%;background:{{ '#2fad5a' if cl=='g' else ('#f59020' if cl=='am' else '#e53935') }}"></div></div></td>
        </tr>
        {% endif %}
      {% endfor %}
      </tbody>
    </table>
  </div>
  <div class="dec-units-grid">
    <div class="dec-unit"><div class="dec-unit-h">💰 Faturamento (VF)</div><div class="dec-unit-b" style="display:grid;grid-template-columns:1fr 1fr;gap:10px">
      <div class="box"><div class="box-lbl">Comercial</div><div class="box-val">{{ res.fat_comercial|brl0 }}</div></div>
      <div class="box"><div class="box-lbl">Matrículas</div><div class="box-val">{{ res.matriculas|toint }}</div></div>
      <div class="box"><div class="box-lbl">Cobr. Atual</div><div class="box-val">{{ res.fin_atual|pct }}</div></div>
      <div class="box"><div class="box-lbl">TPCv1</div><div class="box-val">{{ res.tpcv1|pct }}</div></div>
    </div></div>
    <div class="dec-unit"><div class="dec-unit-h">🚫 Cancelamentos (VF)</div><div class="dec-unit-b" style="display:grid;grid-template-columns:1fr 1fr;gap:10px">
      <div class="box"><div class="box-lbl">Cancelamentos</div><div class="box-val">{{ res.cancelados|toint }}</div></div>
      <div class="box"><div class="box-lbl">Canc. SCPC</div><div class="box-val">{{ res.canc_scpc|toint }}</div></div>
      <div class="box"><div class="box-lbl">Cobr. 30d</div><div class="box-val">{{ res.fin_30|pct }}</div></div>
      <div class="box"><div class="box-lbl">Cobr. 60d</div><div class="box-val">{{ res.fin_60|pct }}</div></div>
    </div></div>
  </div>
</div>
</body></html>
"""

@app.route("/decendial/resultado", methods=["GET"])
def decendial_resultado():
    nome=request.args.get("u","").strip(); d=request.args.get("d","").strip()
    res=_dec["resultados"].get(nome,{}).get(d)
    if not res: return redirect("/decendial/acompanhamento?msg=Resultado+não+encontrado")
    html=DEC_RESULT_TEMPLATE.replace("__CSS__",CSS+CSS_DEC).replace("__NAV__",dec_nav("acomp"))
    return render_template_string(html, unidade=nome, res=res)

DEC_RESULT_PDF = r"""
<!DOCTYPE html><html lang="pt-br"><head><meta charset="UTF-8"><title>IFP – Resultado {{ unidade }}</title>
<link href="https://fonts.googleapis.com/css2?family=Sora:wght@600;700;800&family=Plus+Jakarta+Sans:wght@400;600;700;800&display=swap" rel="stylesheet">
<style>
*{box-sizing:border-box;margin:0;padding:0}body{font-family:'Plus Jakarta Sans',sans-serif;color:#141a36;font-size:11px}
.ph{background:linear-gradient(120deg,#16205e,#1f2d7a);color:#fff;padding:18px 24px;display:flex;align-items:center;justify-content:space-between;margin-bottom:18px}
.ph-logo{width:42px;height:42px;background:#fff;border-radius:10px;display:flex;align-items:center;justify-content:center;font-family:'Sora';font-weight:800;color:#c0021c}
.ph-r h2{font-size:1rem;font-weight:800;text-align:right}.ph-r p{font-size:.68rem;opacity:.85;text-align:right;margin-top:2px}
.no-print{text-align:center;margin-bottom:14px}.no-print button{border:none;border-radius:10px;padding:10px 26px;font-size:.88rem;font-weight:700;cursor:pointer;margin:0 4px}
.btn-p{background:#c0021c;color:#fff}.btn-c{background:#16205e;color:#fff}
table{border-collapse:collapse;width:calc(100% - 36px);margin:0 18px 12px;font-size:11px}
th,td{border:1px solid #e2e6f2;padding:8px 10px;text-align:center}
th{background:#16205e;color:#fff;font-size:.64rem;text-transform:uppercase}
td.lbl{text-align:left;font-weight:700}
tr:nth-child(even){background:#f6f8fc}
.g{color:#1b7a3d;font-weight:800}.am{color:#c2620a;font-weight:800}.r{color:#c0021c;font-weight:800}
.rod{text-align:center;color:#9aa3b2;font-size:.64rem;padding:14px 0;margin:0 18px}
@media print{body{-webkit-print-color-adjust:exact;print-color-adjust:exact}.no-print{display:none}}
</style></head><body>
<div class="ph"><div class="ph-logo">IFP</div><div class="ph-r"><h2>Resultado — {{ unidade }}</h2><p>{{ res.decendio_label }} &nbsp;|&nbsp; Período Decendial</p></div></div>
<div class="no-print"><button class="btn-p" onclick="window.print()">🖨️ Imprimir / Salvar PDF</button><button class="btn-c" onclick="window.close()">✕ Fechar</button></div>
<table>
<tr><th style="text-align:left">Indicador</th><th>Meta do Mês</th><th>Realizado (10d)</th><th>% Atingido</th></tr>
{% for l in res.linhas %}{% if l.meta>0 or l.realizado>0 %}
{% set p=(l.pct*100) %}{% set cl='g' if p>=100 else ('am' if p>=60 else 'r') %}
<tr><td class="lbl">{{ l.label }}</td>
<td>{% if l.tipo=='pct' %}{{ '%.1f%%'|format(l.meta) }}{% else %}{{ l.meta|toint }}{% endif %}</td>
<td>{% if l.tipo=='pct' %}{{ '%.1f%%'|format(l.realizado) }}{% else %}{{ l.realizado|toint }}{% endif %}</td>
<td class="{{ cl }}">{{ '%.0f%%'|format(p) }}</td></tr>
{% endif %}{% endfor %}
</table>
<div class="rod">Gerado pelo sistema IFP Dashboard — Período Decendial</div>
<script>window.addEventListener('load',function(){setTimeout(function(){window.print()},700)});</script>
</body></html>
"""

@app.route("/decendial/resultado/pdf", methods=["GET"])
def decendial_resultado_pdf():
    nome=request.args.get("u","").strip(); d=request.args.get("d","").strip()
    res=_dec["resultados"].get(nome,{}).get(d)
    if not res: return redirect("/decendial/acompanhamento?msg=Resultado+não+encontrado")
    return Response(render_template_string(DEC_RESULT_PDF, unidade=nome, res=res), mimetype="text/html")

if __name__=="__main__":
    port=int(os.environ.get("PORT",5000))
    app.run(host="0.0.0.0",port=port,debug=False)
